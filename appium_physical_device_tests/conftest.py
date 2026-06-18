"""
CrowdSense Appium Physical Device E2E — conftest.py
====================================================
Runs the FULL 100-test E2E suite on a REAL Android phone
connected via USB cable (no emulator required).

How it works — exactly like Selenium auto-opens Chrome:
  1. Plug in your Android phone via USB cable
  2. Enable USB Debugging in Developer Options
  3. Run:  run_physical_tests.bat
  4. Appium auto-installs the APK, opens the app on your phone,
     taps every screen, and generates an Excel report.

Environment Variables (optional — override defaults):
    CROWDSENSE_EMAIL      : Firebase login email    (default: test@crowdsense.app)
    CROWDSENSE_PASSWORD   : Firebase login password (default: Test@1234)
    APPIUM_HOST           : Appium server host      (default: 127.0.0.1)
    APPIUM_PORT           : Appium server port      (default: 4723)
    APK_PATH              : Path to debug APK
                            (default: ../frontend/build/app/outputs/flutter-apk/app-debug.apk)
    ANDROID_DEVICE_ID     : Force a specific device serial (optional — auto-detected if blank)
"""

import os
import subprocess
import time
import datetime
import pytest
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from appium import webdriver as appium_webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# ── Configuration ──────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_APK = os.path.join(
    _HERE, "..", "frontend", "build", "app", "outputs",
    "flutter-apk", "app-debug.apk"
)

APPIUM_HOST   = os.environ.get("APPIUM_HOST",         "127.0.0.1")
APPIUM_PORT   = int(os.environ.get("APPIUM_PORT",     "4723"))
APK_PATH      = os.path.abspath(os.environ.get("APK_PATH", _DEFAULT_APK))
TEST_EMAIL    = os.environ.get("CROWDSENSE_EMAIL",    "test@crowdsense.app")
TEST_PASSWORD = os.environ.get("CROWDSENSE_PASSWORD", "Test@1234")
FORCED_DEVICE = os.environ.get("ANDROID_DEVICE_ID",  "")

APPIUM_SERVER = f"http://{APPIUM_HOST}:{APPIUM_PORT}"
REPORTS_DIR   = os.path.join(_HERE, "reports")

# ── Excel style constants ───────────────────────────────────────────────────────
HDR_FILL    = PatternFill("solid", fgColor="1A3A5C")   # dark navy
PASS_FILL   = PatternFill("solid", fgColor="C6EFCE")   # light green
FAIL_FILL   = PatternFill("solid", fgColor="FFC7CE")   # light red
SKIP_FILL   = PatternFill("solid", fgColor="FFEB9C")   # amber
WHITE_FONT  = Font(name="Calibri", color="FFFFFF", bold=True)
BOLD_FONT   = Font(name="Calibri", bold=True)
NORMAL_FONT = Font(name="Calibri")
thin        = Side(style="thin", color="B0B0B0")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Global result store ─────────────────────────────────────────────────────────
_suite_start: datetime.datetime = None
_results: list = []
_device_info: dict = {}   # populated at session start with real device details


# ─────────────────────────────────────────────────────────────────────────────
# USB Device Auto-Detection (the key difference vs emulator tests)
# ─────────────────────────────────────────────────────────────────────────────

def _detect_usb_device() -> str:
    """
    Uses 'adb devices' to auto-detect the first real USB-connected
    Android device. Returns the device serial ID, or '' if none found.
    Physical devices appear as  <serial>  device  (not 'emulator-...')
    """
    try:
        result = subprocess.run(
            ["adb", "devices"],
            capture_output=True, text=True, timeout=10
        )
        lines = result.stdout.strip().splitlines()
        for line in lines[1:]:  # skip header "List of devices attached"
            parts = line.strip().split()
            if len(parts) >= 2 and parts[1] == "device":
                serial = parts[0]
                # Skip emulators — we want a real USB device
                if not serial.startswith("emulator-"):
                    print(f"\n[DEVICE] Auto-detected USB device: {serial}")
                    return serial
    except (FileNotFoundError, subprocess.TimeoutExpired) as e:
        print(f"\n[WARN] ADB not found or timed out: {e}")
    return ""


def _get_device_info(serial: str) -> dict:
    """
    Reads device brand, model, Android version from ADB for the report header.
    """
    info = {"serial": serial, "brand": "Unknown", "model": "Unknown", "android": "Unknown"}
    if not serial:
        return info

    def _adb_prop(prop):
        try:
            r = subprocess.run(
                ["adb", "-s", serial, "shell", "getprop", prop],
                capture_output=True, text=True, timeout=8
            )
            return r.stdout.strip() or "Unknown"
        except Exception:
            return "Unknown"

    info["brand"]   = _adb_prop("ro.product.brand").title()
    info["model"]   = _adb_prop("ro.product.model")
    info["android"] = _adb_prop("ro.build.version.release")
    info["sdk"]     = _adb_prop("ro.build.version.sdk")
    return info


# ─────────────────────────────────────────────────────────────────────────────
# pytest hooks
# ─────────────────────────────────────────────────────────────────────────────

def pytest_sessionstart(session):
    global _suite_start, _device_info
    _suite_start = datetime.datetime.now(datetime.timezone.utc)

    # Detect USB device early so failures appear before tests run
    serial = FORCED_DEVICE or _detect_usb_device()
    if serial:
        _device_info = _get_device_info(serial)
        print(f"\n[DEVICE] {_device_info['brand']} {_device_info['model']} "
              f"(Android {_device_info['android']}, SDK {_device_info.get('sdk','?')}) "
              f"- Serial: {serial}")
        # Clear app cache and databases to prevent Android from restoring session via auto-backup
        try:
            print(f"[DEVICE] Resetting app state via ADB on {serial}...")
            subprocess.run(["adb", "-s", serial, "shell", "pm", "clear", "com.example.crowdsense"], capture_output=True, timeout=8)
        except Exception as e:
            print(f"[WARN] Failed to clear app state: {e}")

    else:
        print("\n[WARN] No USB device detected. Tests will likely fail.")
        _device_info = {"serial": "none", "brand": "Unknown", "model": "Unknown",
                        "android": "Unknown", "sdk": "Unknown"}


def _get_category(nodeid):
    """Extract folder name as category."""
    parts = nodeid.split(os.sep)
    return parts[0] if len(parts) > 1 else "root"


def pytest_runtest_logreport(report):
    global _results
    if report.when == "call":
        status = "PASSED" if report.passed else ("FAILED" if report.failed else "SKIPPED")
        error  = (
            str(report.longrepr)[:800]
            if report.failed
            else "None — test passed successfully."
        )
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _results.append({
            "no":       len(_results) + 1,
            "category": _get_category(report.nodeid),
            "name":     report.nodeid.split("::")[-1],
            "status":   status,
            "duration": round(report.duration, 2),
            "error":    error,
            "ts":       ts,
        })


def _write_header(ws, columns, row=1):
    for col_idx, hdr in enumerate(columns, start=1):
        c = ws.cell(row=row, column=col_idx, value=hdr)
        c.font = WHITE_FONT; c.fill = HDR_FILL; c.border = THIN_BORDER; c.alignment = CENTER
    ws.row_dimensions[row].height = 26


def _generate_xlsx():
    if not _results:
        print("\n[REPORT] No test results to write.")
        return

    suite_end = datetime.datetime.now(datetime.timezone.utc)
    suite_start_str = _suite_start.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z" if _suite_start else ""
    suite_end_str = suite_end.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"

    passed_rows = [r for r in _results if r["status"] == "PASSED"]
    failed_rows = [r for r in _results if r["status"] == "FAILED"]
    skipped_rows = [r for r in _results if r["status"] == "SKIPPED"]

    total = len(_results)
    passed_count = len(passed_rows)
    failed_count = len(failed_rows)
    skipped_count = len(skipped_rows)
    pass_rate = round(passed_count / total * 100, 1) if total else 0.0
    duration_sec = sum(r["duration"] for r in _results)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    # Row 1 Title
    ws_sum.merge_cells("A1:H1")
    t = ws_sum["A1"]
    t.value = "CrowdSense — Appium Mobile E2E Test Report" if "emulator" in _device_info.get("serial", "") else "CrowdSense — Physical Device Appium E2E Test Report"
    t.font = Font(name="Calibri", color="FFFFFF", bold=True, size=16)
    t.fill = PatternFill("solid", fgColor="1A3A5C")
    t.alignment = CENTER
    ws_sum.row_dimensions[1].height = 42.0

    # Row 2 Subtitle
    ws_sum.merge_cells("A2:H2")
    sub = ws_sum["A2"]
    sub.value = (
        f"Device: {_device_info.get('brand','Unknown')} {_device_info.get('model','Unknown')} "
        f"(Android {_device_info.get('android','Unknown')}, Serial: {_device_info.get('serial','Unknown')})  |  "
        f"Automation: UiAutomator2 (USB)  |  Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    sub.font = Font(name="Calibri", color="FFFFFF", size=10)
    sub.fill = PatternFill("solid", fgColor="2D5F8A")
    sub.alignment = CENTER
    ws_sum.row_dimensions[2].height = 20.0
    ws_sum.row_dimensions[3].height = 8.0

    # Row 4 Table headers
    _write_header(ws_sum, ["Test Suite", "Total", "Passed", "Failed", "Skipped", "Pass Rate %", "Duration (s)", "Start Time"], row=4)
    ws_sum.row_dimensions[4].height = 26.0

    # Row 5 Table values
    suite_title = "CrowdSense Android — Full Appium E2E (Emulator)" if "emulator" in _device_info.get("serial", "") else "CrowdSense Android — Full Appium E2E (Physical Device)"
    row_vals = [
        suite_title,
        total,
        passed_count,
        failed_count,
        skipped_count,
        f"{pass_rate}%",
        round(duration_sec, 2),
        suite_start_str
    ]
    for col_idx, val in enumerate(row_vals, start=1):
        c = ws_sum.cell(row=5, column=col_idx, value=val)
        c.font = NORMAL_FONT
        c.border = THIN_BORDER
        c.alignment = LEFT if col_idx == 1 else CENTER
    ws_sum.row_dimensions[5].height = 20.0
    ws_sum.row_dimensions[6].height = 8.0

    # Row 7 Device Info headers
    _write_header(ws_sum, ["Device Info", "Value"], row=7)
    # Clear out cells C7:H7 just in case they have borders from _write_header
    for col in range(3, 9):
        ws_sum.cell(row=7, column=col).border = Border()
    ws_sum.row_dimensions[7].height = 26.0

    # Device Info values
    device_details = [
        ("Brand", _device_info.get("brand", "Unknown")),
        ("Model", _device_info.get("model", "Unknown")),
        ("Android Version", _device_info.get("android", "Unknown")),
        ("SDK Level", _device_info.get("sdk", "Unknown")),
        ("Serial / ID", _device_info.get("serial", "Unknown"))
    ]
    for idx, (prop, val) in enumerate(device_details, start=8):
        c_prop = ws_sum.cell(row=idx, column=1, value=prop)
        c_val = ws_sum.cell(row=idx, column=2, value=val)
        c_prop.font = BOLD_FONT
        c_prop.border = THIN_BORDER
        c_prop.alignment = LEFT
        c_val.font = NORMAL_FONT
        c_val.border = THIN_BORDER
        c_val.alignment = LEFT
        ws_sum.row_dimensions[idx].height = 20.0

    # Column widths
    ws_sum.column_dimensions["A"].width = 55.0
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws_sum.column_dimensions[col].width = 14.0
    ws_sum.column_dimensions["H"].width = 26.0

    # ── Sheet 2: Passed Tests ────────────────────────────────────────────────
    ws_pass = wb.create_sheet("Passed Tests")
    _write_header(ws_pass, ["No.", "Category", "Test Name", "Duration (s)", "Status"], row=1)
    for r_idx, r in enumerate(passed_rows, start=2):
        vals = [r["no"], r["category"], r["name"], r["duration"], r["status"]]
        for c_idx, val in enumerate(vals, start=1):
            c = ws_pass.cell(row=r_idx, column=c_idx, value=val)
            c.font = NORMAL_FONT
            c.fill = PASS_FILL
            c.border = THIN_BORDER
            c.alignment = LEFT if c_idx in (2, 3) else CENTER
        ws_pass.row_dimensions[r_idx].height = 20.0
    ws_pass.column_dimensions["A"].width = 6.0
    ws_pass.column_dimensions["B"].width = 30.0
    ws_pass.column_dimensions["C"].width = 72.0
    ws_pass.column_dimensions["D"].width = 14.0
    ws_pass.column_dimensions["E"].width = 12.0

    # ── Sheet 3: Failed Tests ────────────────────────────────────────────────
    ws_fail = wb.create_sheet("Failed Tests")
    _write_header(ws_fail, ["No.", "Category", "Test Name", "Error / Traceback", "Status", "Timestamp"], row=1)
    for r_idx, r in enumerate(failed_rows, start=2):
        vals = [r["no"], r["category"], r["name"], r["error"], r["status"], r["ts"]]
        for c_idx, val in enumerate(vals, start=1):
            c = ws_fail.cell(row=r_idx, column=c_idx, value=val)
            c.font = NORMAL_FONT
            c.fill = FAIL_FILL
            c.border = THIN_BORDER
            c.alignment = LEFT if c_idx in (2, 3, 4) else CENTER
        ws_fail.row_dimensions[r_idx].height = 20.0
    ws_fail.column_dimensions["A"].width = 6.0
    ws_fail.column_dimensions["B"].width = 28.0
    ws_fail.column_dimensions["C"].width = 55.0
    ws_fail.column_dimensions["D"].width = 90.0
    ws_fail.column_dimensions["E"].width = 12.0
    ws_fail.column_dimensions["F"].width = 22.0

    # ── Sheet 4: Execution Log ───────────────────────────────────────────────
    ws_log = wb.create_sheet("Execution Log")
    _write_header(ws_log, ["Timestamp", "Level", "Category", "Test Name", "Result", "Duration (s)"], row=1)
    for r_idx, r in enumerate(_results, start=2):
        lvl = "INFO" if r["status"] == "PASSED" else "ERROR"
        vals = [r["ts"], lvl, r["category"], r["name"], r["status"], r["duration"]]
        fill = PatternFill("solid", fgColor="C6EFCE") if r["status"] == "PASSED" else (PatternFill("solid", fgColor="FFC7CE") if r["status"] == "FAILED" else PatternFill("solid", fgColor="FFEB9C"))
        for c_idx, val in enumerate(vals, start=1):
            c = ws_log.cell(row=r_idx, column=c_idx, value=val)
            c.font = NORMAL_FONT
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = LEFT if c_idx in (3, 4) else CENTER
        ws_log.row_dimensions[r_idx].height = 20.0
    ws_log.column_dimensions["A"].width = 22.0
    ws_log.column_dimensions["B"].width = 8.0
    ws_log.column_dimensions["C"].width = 28.0
    ws_log.column_dimensions["D"].width = 65.0
    ws_log.column_dimensions["E"].width = 12.0
    ws_log.column_dimensions["F"].width = 14.0

    # ── Sheet 5: Test Details ────────────────────────────────────────────────
    ws_det = wb.create_sheet("Test Details")
    _write_header(ws_det, ["No.", "Category", "Test Name", "Status", "Duration (s)", "Error Details"], row=1)
    for r_idx, r in enumerate(_results, start=2):
        vals = [r["no"], r["category"], r["name"], r["status"], r["duration"], r["error"]]
        fill = PASS_FILL if r["status"] == "PASSED" else (FAIL_FILL if r["status"] == "FAILED" else SKIP_FILL)
        for c_idx, val in enumerate(vals, start=1):
            c = ws_det.cell(row=r_idx, column=c_idx, value=val)
            c.font = NORMAL_FONT
            c.fill = fill
            c.border = THIN_BORDER
            c.alignment = LEFT if c_idx in (2, 3, 6) else CENTER
        ws_det.row_dimensions[r_idx].height = 22.0
    ws_det.column_dimensions["A"].width = 6.0
    ws_det.column_dimensions["B"].width = 28.0
    ws_det.column_dimensions["C"].width = 60.0
    ws_det.column_dimensions["D"].width = 12.0
    ws_det.column_dimensions["E"].width = 14.0
    ws_det.column_dimensions["F"].width = 90.0
    ws_det.freeze_panes = "A2"

    ts = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    prefix = "Physical_Device_E2E_Report" if "emulator" not in _device_info.get("serial", "") else "Appium_E2E_Report_CrowdSense"
    fname = os.path.join(REPORTS_DIR, f"{prefix}_{ts}.xlsx")
    wb.save(fname)
    print(f"\n[REPORT] Appium 5-sheet report saved -> {fname}")


# ─────────────────────────────────────────────────────────────────────────────
# Appium Driver Fixtures  — Physical Device (USB)
# ─────────────────────────────────────────────────────────────────────────────

def _build_physical_options(device_serial: str) -> UiAutomator2Options:
    """
    Build Appium UiAutomator2 capabilities for a REAL USB device.
    Key differences from emulator config:
      - NO 'avd' capability (no emulator to launch)
      - device_name = the ADB serial of the physical phone
      - udid = same serial (tells Appium exactly which USB device to use)
      - no_reset = True (keeps existing app data — avoids re-install on every test)
    """
    options = UiAutomator2Options()
    options.platform_name       = "Android"
    options.device_name         = device_serial        # Real device serial
    options.udid                = device_serial        # Explicit USB device ID
    options.app                 = APK_PATH
    options.app_package         = "com.example.crowdsense"
    options.app_activity        = "com.example.crowdsense.MainActivity"
    options.automation_name     = "UiAutomator2"
    options.no_reset            = False                # Fresh install
    options.full_reset          = False
    options.new_command_timeout = 180                  # 3 min (real devices can be slower)
    options.android_install_timeout = 120000           # ms

    # Flutter / UiAutomator2 performance settings
    options.set_capability("settings[waitForIdleTimeout]",    10)
    options.set_capability("settings[waitForSelectorTimeout]",10000)
    options.set_capability("disableWindowAnimation",          True)

    # Skip animations on real device for faster, stable tests
    options.set_capability("ignoreUnimportantViews", True)

    return options


@pytest.fixture(scope="session")
def driver():
    """
    Session-scoped Appium driver fixture for a REAL USB-connected Android device.

    On session start:
      1. auto-detects the USB device via ADB
      2. connects to Appium server (must be running: npx appium)
      3. installs the Flutter debug APK directly onto the phone
      4. launches the CrowdSense app — you will SEE it open on your phone
      5. yields the driver for the full test session
      6. quits / unlocks the phone when all tests finish
    """
    # ── Validate APK exists ───────────────────────────────────────────────────
    if not os.path.isfile(APK_PATH):
        pytest.skip(
            f"[SKIP] APK not found at: {APK_PATH}\n"
            "Build it first:  cd frontend && flutter build apk --debug"
        )

    # ── Detect USB device ─────────────────────────────────────────────────────
    serial = FORCED_DEVICE or _detect_usb_device()
    if not serial:
        pytest.skip(
            "[SKIP] No USB Android device detected.\n"
            "Connect your phone via USB cable and enable USB Debugging.\n"
            "Run 'adb devices' to verify your phone is listed."
        )

    # ── Connect Appium ────────────────────────────────────────────────────────
    options = _build_physical_options(serial)
    print(f"\n[APPIUM] Connecting to {APPIUM_SERVER} -> installing APK on {serial}...")
    print("[APPIUM] Your phone screen will light up and the app will open automatically.")

    drv = appium_webdriver.Remote(APPIUM_SERVER, options=options)
    drv.implicitly_wait(0)   # Use explicit waits instead to avoid mixing wait types
    yield drv
    drv.quit()


@pytest.fixture(scope="module")
def logged_in_driver(driver):
    """
    Module-scoped fixture: performs Firebase login once per test module.
    Returns the driver already on the Home screen.
    """
    _perform_login(driver)
    yield driver


def _ensure_home_conftest(drv):
    """Ensure the app is on the Home/Trending screen. If stuck on a sub-screen, restart the app."""
    # Check if already home/logged in
    for xpath in [
        '//*[contains(@content-desc,"Trending") or contains(@text,"Trending")]',
    ]:
        try:
            if drv.find_elements(AppiumBy.XPATH, xpath):
                return True
        except Exception:
            pass

    # Try tapping Home tab in bottom nav
    try:
        home_tab = drv.find_element(AppiumBy.XPATH, '//*[contains(@content-desc,"Home") or contains(@text,"Home")]')
        if home_tab:
            home_tab.click()
            time.sleep(2.5)
            if drv.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Trending") or contains(@text,"Trending")]'):
                return True
    except Exception:
        pass

    try:
        # Check for Flutter multiline content description tab names
        if drv.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Home\n") or contains(@text,"Home\n")]'):
            home_tab = drv.find_element(AppiumBy.XPATH, '//*[contains(@content-desc,"Home\n") or contains(@text,"Home\n")]')
            home_tab.click()
            time.sleep(2.5)
            if drv.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Trending") or contains(@text,"Trending")]'):
                return True
    except Exception:
        pass

    # Try back button up to 4 times
    for _ in range(4):
        try:
            if drv.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Trending") or contains(@text,"Trending")]'):
                return True
            drv.back()
            time.sleep(2)
        except Exception:
            break

    # Last resort: restart app to clean session state
    try:
        print("[INFO] Resetting app by restarting...")
        drv.terminate_app("com.example.crowdsense")
        time.sleep(2)
        drv.activate_app("com.example.crowdsense")
        
        # Wait up to 15 seconds for splash screen to transition to either home or login screen
        WebDriverWait(drv, 15).until(
            EC.presence_of_element_located((AppiumBy.XPATH, 
                '//*[contains(@content-desc,"Trending") or contains(@text,"Trending") or contains(@content-desc,"Sign In") or contains(@text,"Sign In")]'
            ))
        )
        if drv.find_elements(AppiumBy.XPATH, '//*[contains(@content-desc,"Trending") or contains(@text,"Trending")]'):
            return True
    except Exception:
        pass

    return False


def _perform_login(drv):
    """Helper: fill email + password + tap Sign In on the login screen."""
    # Ensure app is on home screen if already logged in, otherwise recover if stuck on sub-screen
    if _ensure_home_conftest(drv):
        print("[INFO] App is already logged in and on Home screen.")
        return

    # Check if we are on Login screen (or wait for it to load)
    try:
        WebDriverWait(drv, 15).until(
            EC.presence_of_element_located((AppiumBy.XPATH, 
                '//*[@content-desc="Email" or @text="Email"] | //*[@content-desc="Sign In" or @text="Sign In"] | //android.widget.EditText'
            ))
        )
    except Exception as e:
        print(f"[WARN] Timeout waiting for Login screen to load: {e}")

    try:
        # Give the UI 2 seconds to settle
        time.sleep(2)

        # Locate Email Field
        email_field = None

        for xpath in [
            '//android.widget.EditText[1]',
            '//*[@content-desc="Email" or @text="Email"]',
            '//*[contains(@content-desc,"email") or contains(@text,"email")]',
            '//*[contains(@resource-id,"email")]'
        ]:
            try:
                email_field = drv.find_element(AppiumBy.XPATH, xpath)
                if email_field:
                    break
            except Exception:
                continue

        if not email_field:
            raise NoSuchElementException("Could not locate Email input field using any locator.")

        email_field.click()
        time.sleep(1)
        email_field.clear()
        email_field.send_keys(TEST_EMAIL)
        time.sleep(1)

        # Locate Password Field
        pwd_field = None
        for xpath in [
            '//android.widget.EditText[2]',
            '//*[@content-desc="Password" or @text="Password"]',
            '//*[contains(@content-desc,"password") or contains(@text,"password")]',
            '//*[contains(@resource-id,"password")]'
        ]:
            try:
                pwd_field = drv.find_element(AppiumBy.XPATH, xpath)
                if pwd_field:
                    break
            except Exception:
                continue

        if not pwd_field:
            raise NoSuchElementException("Could not locate Password input field using any locator.")

        pwd_field.click()
        time.sleep(1)
        pwd_field.clear()
        pwd_field.send_keys(TEST_PASSWORD)
        time.sleep(1)

        # Hide keyboard so it doesn't block the Sign In button
        try:
            drv.hide_keyboard()
            time.sleep(1)
        except Exception:
            pass

        # Locate Sign In Button
        sign_in_btn = None
        for xpath in [
            '//*[@content-desc="Sign In" or @text="Sign In"]',
            '//*[contains(@content-desc,"Sign") or contains(@text,"Sign")]',
            '//*[contains(@resource-id,"sign_in")]'
        ]:
            try:
                sign_in_btn = drv.find_element(AppiumBy.XPATH, xpath)
                if sign_in_btn:
                    break
            except Exception:
                continue

        if not sign_in_btn:
            raise NoSuchElementException("Could not locate Sign In button using any locator.")

        sign_in_btn.click()
        time.sleep(10)   # Wait for Firebase authentication and transition to Home screen
    except Exception as e:
        print(f"[WARN] Login helper failed (may already be logged in): {e}")
        raise e


@pytest.fixture
def app_package():
    return "com.example.crowdsense"


@pytest.fixture
def appium_server():
    return APPIUM_SERVER
