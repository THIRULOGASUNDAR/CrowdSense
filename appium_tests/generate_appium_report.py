"""
CrowdSense — Standalone Appium Mobile E2E XLSX Report Generator
===============================================================
Run this script BEFORE or AFTER running pytest to generate a pre-filled
Excel report template listing all 100 Appium test cases (AM001–AM100).

Usage:
    python generate_appium_report.py
    python generate_appium_report.py my_output.xlsx
"""
import datetime
import os
import sys
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── All 100 Appium test cases ────────────────────────────────────────────────
TEST_CASES = [
    # Authentication (AM001–AM022)
    ("AM001", "test_01_auth", "App launches and shows splash screen",               "Authentication", "P1 – Critical"),
    ("AM002", "test_01_auth", "Splash auto-redirects to Login within 5 seconds",    "Authentication", "P1 – Critical"),
    ("AM003", "test_01_auth", "Login screen has email input field",                 "Authentication", "P1 – Critical"),
    ("AM004", "test_01_auth", "Login screen has password input field",              "Authentication", "P1 – Critical"),
    ("AM005", "test_01_auth", "Login screen has Sign In button",                    "Authentication", "P1 – Critical"),
    ("AM006", "test_01_auth", "Login screen shows Welcome Back heading",            "Authentication", "P2 – High"),
    ("AM007", "test_01_auth", "Login screen shows Forgot Password link",            "Authentication", "P2 – High"),
    ("AM008", "test_01_auth", "Login screen shows Sign Up link",                    "Authentication", "P2 – High"),
    ("AM009", "test_01_auth", "Empty login form shows validation message",          "Authentication", "P2 – High"),
    ("AM010", "test_01_auth", "Wrong credentials keeps user on Login screen",       "Authentication", "P2 – High"),
    ("AM011", "test_01_auth", "Navigate to Register page via Sign Up link",         "Authentication", "P1 – Critical"),
    ("AM012", "test_01_auth", "Register page has Full Name input field",            "Authentication", "P2 – High"),
    ("AM013", "test_01_auth", "Register page has Email input field",                "Authentication", "P2 – High"),
    ("AM014", "test_01_auth", "Register page has Password input field",             "Authentication", "P2 – High"),
    ("AM015", "test_01_auth", "Register page has Confirm Password field",           "Authentication", "P2 – High"),
    ("AM016", "test_01_auth", "Register page has Create Account button",            "Authentication", "P2 – High"),
    ("AM017", "test_01_auth", "Password mismatch shows validation error",           "Authentication", "P2 – High"),
    ("AM018", "test_01_auth", "Forgot Password page renders",                       "Authentication", "P2 – High"),
    ("AM019", "test_01_auth", "Forgot Password has email input field",              "Authentication", "P3 – Medium"),
    ("AM020", "test_01_auth", "Forgot Password has Send Reset Link button",         "Authentication", "P3 – Medium"),
    ("AM021", "test_01_auth", "Route guard: unauthenticated user redirected to Login", "Authentication", "P1 – Critical"),
    ("AM022", "test_01_auth", "Successful login navigates to Home screen",          "Authentication", "P1 – Critical"),
    # Home & Navigation (AM023–AM040)
    ("AM023", "test_02_home_navigation", "Home screen loads after login",                    "Home & Navigation", "P1 – Critical"),
    ("AM024", "test_02_home_navigation", "Home shows search bar",                           "Home & Navigation", "P1 – Critical"),
    ("AM025", "test_02_home_navigation", "Home shows greeting / welcome text",              "Home & Navigation", "P2 – High"),
    ("AM026", "test_02_home_navigation", "Home shows Trending Now section",                 "Home & Navigation", "P2 – High"),
    ("AM027", "test_02_home_navigation", "Home shows Popular Categories section",           "Home & Navigation", "P2 – High"),
    ("AM028", "test_02_home_navigation", "Landmarks category chip visible on Home",         "Home & Navigation", "P3 – Medium"),
    ("AM029", "test_02_home_navigation", "Restaurants category chip visible on Home",       "Home & Navigation", "P3 – Medium"),
    ("AM030", "test_02_home_navigation", "Parks category chip visible on Home",             "Home & Navigation", "P3 – Medium"),
    ("AM031", "test_02_home_navigation", "Shopping category chip visible on Home",          "Home & Navigation", "P3 – Medium"),
    ("AM032", "test_02_home_navigation", "Entertainment category chip visible on Home",     "Home & Navigation", "P3 – Medium"),
    ("AM033", "test_02_home_navigation", "Home shows Recent Searches section",              "Home & Navigation", "P3 – Medium"),
    ("AM034", "test_02_home_navigation", "Bottom navigation bar visible",                   "Home & Navigation", "P1 – Critical"),
    ("AM035", "test_02_home_navigation", "Tap Search tab navigates to Search screen",       "Home & Navigation", "P1 – Critical"),
    ("AM036", "test_02_home_navigation", "Tap Planner tab navigates to Travel Planner",     "Home & Navigation", "P2 – High"),
    ("AM037", "test_02_home_navigation", "Tap Favorites tab navigates to Favorites screen", "Home & Navigation", "P2 – High"),
    ("AM038", "test_02_home_navigation", "Tap Profile tab navigates to Profile screen",     "Home & Navigation", "P2 – High"),
    ("AM039", "test_02_home_navigation", "Back navigation from Search returns to Home",     "Home & Navigation", "P2 – High"),
    ("AM040", "test_02_home_navigation", "All bottom navigation icons render",              "Home & Navigation", "P2 – High"),
    # Search & Place (AM041–AM060)
    ("AM041", "test_03_search_place", "Search screen loads with text input",                   "Search & Place", "P1 – Critical"),
    ("AM042", "test_03_search_place", "Search shows suggestion chips (Coffee Shop, Library)",  "Search & Place", "P2 – High"),
    ("AM043", "test_03_search_place", "Search screen shows Recent Searches section",           "Search & Place", "P2 – High"),
    ("AM044", "test_03_search_place", "Typing in search box triggers results",                 "Search & Place", "P1 – Critical"),
    ("AM045", "test_03_search_place", "Search results list renders",                           "Search & Place", "P1 – Critical"),
    ("AM046", "test_03_search_place", "Tap search result navigates to Place Details",          "Search & Place", "P1 – Critical"),
    ("AM047", "test_03_search_place", "Place Details shows place name",                        "Search & Place", "P1 – Critical"),
    ("AM048", "test_03_search_place", "Place Details shows location / address info",           "Search & Place", "P2 – High"),
    ("AM049", "test_03_search_place", "Place Details shows current crowd status card",         "Search & Place", "P1 – Critical"),
    ("AM050", "test_03_search_place", "Place Details shows Best Time to Visit section",        "Search & Place", "P2 – High"),
    ("AM051", "test_03_search_place", "Place Details shows Crowd Forecast section",            "Search & Place", "P2 – High"),
    ("AM052", "test_03_search_place", "Place Details has Report Live Crowd button",            "Search & Place", "P2 – High"),
    ("AM053", "test_03_search_place", "Place Details shows About section",                     "Search & Place", "P2 – High"),
    ("AM054", "test_03_search_place", "Favorite (heart) icon visible in Place Details",        "Search & Place", "P2 – High"),
    ("AM055", "test_03_search_place", "Share icon visible in Place Details app bar",           "Search & Place", "P3 – Medium"),
    ("AM056", "test_03_search_place", "Crowd Report has Not Busy / Low option",                "Search & Place", "P2 – High"),
    ("AM057", "test_03_search_place", "Crowd Report has A Bit Busy / Moderate option",        "Search & Place", "P2 – High"),
    ("AM058", "test_03_search_place", "Crowd Report has Very Crowded / High option",           "Search & Place", "P2 – High"),
    ("AM059", "test_03_search_place", "Crowd Report has optional note / comment field",        "Search & Place", "P3 – Medium"),
    ("AM060", "test_03_search_place", "Community Photos page loads",                           "Search & Place", "P3 – Medium"),
    # Profile & Settings (AM061–AM078)
    ("AM061", "test_04_profile_settings", "Profile screen renders",                        "Profile & Settings", "P1 – Critical"),
    ("AM062", "test_04_profile_settings", "Profile screen shows user name",                "Profile & Settings", "P2 – High"),
    ("AM063", "test_04_profile_settings", "Profile shows Photos stat counter",             "Profile & Settings", "P2 – High"),
    ("AM064", "test_04_profile_settings", "Profile shows Reports stat counter",            "Profile & Settings", "P2 – High"),
    ("AM065", "test_04_profile_settings", "Profile shows Saved (favorites) stat counter", "Profile & Settings", "P2 – High"),
    ("AM066", "test_04_profile_settings", "Edit Profile option visible in Profile",        "Profile & Settings", "P2 – High"),
    ("AM067", "test_04_profile_settings", "My Reports option visible in Profile",          "Profile & Settings", "P2 – High"),
    ("AM068", "test_04_profile_settings", "Notifications option visible in Profile",       "Profile & Settings", "P3 – Medium"),
    ("AM069", "test_04_profile_settings", "Support and FAQ option visible in Profile",     "Profile & Settings", "P3 – Medium"),
    ("AM070", "test_04_profile_settings", "Sign Out button visible in Profile",            "Profile & Settings", "P1 – Critical"),
    ("AM071", "test_04_profile_settings", "My Reports screen loads and renders",           "Profile & Settings", "P2 – High"),
    ("AM072", "test_04_profile_settings", "Settings screen renders",                       "Profile & Settings", "P1 – Critical"),
    ("AM073", "test_04_profile_settings", "Settings has Dark Mode toggle switch",          "Profile & Settings", "P2 – High"),
    ("AM074", "test_04_profile_settings", "Settings has Notifications toggle switch",      "Profile & Settings", "P2 – High"),
    ("AM075", "test_04_profile_settings", "Settings shows Privacy Policy option",          "Profile & Settings", "P3 – Medium"),
    ("AM076", "test_04_profile_settings", "Settings shows Terms of Service option",        "Profile & Settings", "P3 – Medium"),
    ("AM077", "test_04_profile_settings", "Settings shows App Version 1.0.0",             "Profile & Settings", "P3 – Medium"),
    ("AM078", "test_04_profile_settings", "Settings has Delete Account option",            "Profile & Settings", "P2 – High"),
    # Travel Planner & Favorites (AM079–AM090)
    ("AM079", "test_05_travel_planner", "Travel Planner screen renders",                   "Travel & Favorites", "P1 – Critical"),
    ("AM080", "test_05_travel_planner", "Travel Planner shows heading / title text",       "Travel & Favorites", "P2 – High"),
    ("AM081", "test_05_travel_planner", "Travel Planner has From location selector",       "Travel & Favorites", "P2 – High"),
    ("AM082", "test_05_travel_planner", "Travel Planner has To destination selector",      "Travel & Favorites", "P2 – High"),
    ("AM083", "test_05_travel_planner", "Travel Planner has Calculate Best Plan button",   "Travel & Favorites", "P2 – High"),
    ("AM084", "test_05_travel_planner", "Travel Planner shows crowd-aware description",    "Travel & Favorites", "P3 – Medium"),
    ("AM085", "test_05_travel_planner", "Favorites screen renders",                        "Travel & Favorites", "P1 – Critical"),
    ("AM086", "test_05_travel_planner", "Favorites shows empty state or list items",       "Travel & Favorites", "P2 – High"),
    ("AM087", "test_05_travel_planner", "Add place to favorites works",                    "Travel & Favorites", "P2 – High"),
    ("AM088", "test_05_travel_planner", "Favorite heart icon toggles without crash",       "Travel & Favorites", "P2 – High"),
    ("AM089", "test_05_travel_planner", "Unfavorite action updates the Favorites list",    "Travel & Favorites", "P3 – Medium"),
    ("AM090", "test_05_travel_planner", "Favorites list updates after adding a place",     "Travel & Favorites", "P3 – Medium"),
    # Smoke Tests (AM091–AM100)
    ("AM091", "test_06_smoke", "App launches without crash",                        "Smoke Tests", "P1 – Critical"),
    ("AM092", "test_06_smoke", "Login screen accessible on fresh launch",           "Smoke Tests", "P1 – Critical"),
    ("AM093", "test_06_smoke", "Register screen accessible from Login",             "Smoke Tests", "P1 – Critical"),
    ("AM094", "test_06_smoke", "Forgot Password screen accessible",                 "Smoke Tests", "P1 – Critical"),
    ("AM095", "test_06_smoke", "App renders meaningful content within 10 seconds",  "Smoke Tests", "P1 – Critical"),
    ("AM096", "test_06_smoke", "Full Login → Home flow completes",                  "Smoke Tests", "P1 – Critical"),
    ("AM097", "test_06_smoke", "Home → Search → Back flow completes",               "Smoke Tests", "P1 – Critical"),
    ("AM098", "test_06_smoke", "Profile screen accessible after login",             "Smoke Tests", "P1 – Critical"),
    ("AM099", "test_06_smoke", "Settings screen accessible after login",            "Smoke Tests", "P1 – Critical"),
    ("AM100", "test_06_smoke", "App handles rapid screen transitions without crash","Smoke Tests", "P1 – Critical"),
]


def generate_xlsx_report(output_path=None):
    """Generate a styled XLSX Appium test report template with all 100 test cases."""
    os.makedirs("reports", exist_ok=True)
    if output_path is None:
        ts = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
        output_path = os.path.join("reports", f"Appium_E2E_Report_CrowdSense_{ts}.xlsx")

    wb = openpyxl.Workbook()

    # ── Style constants ──────────────────────────────────────────────────────
    PASS_FILL    = PatternFill("solid", fgColor="C6EFCE")
    FAIL_FILL    = PatternFill("solid", fgColor="FFC7CE")
    PENDING_FILL = PatternFill("solid", fgColor="FFEB9C")
    HEADER_FILL  = PatternFill("solid", fgColor="2D5F8A")
    TITLE_FILL   = PatternFill("solid", fgColor="1A3A5C")
    ALT_FILL     = PatternFill("solid", fgColor="EBF2FA")

    WHITE_FONT   = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    BOLD_FONT    = Font(name="Calibri", bold=True, size=11)
    NORMAL_FONT  = Font(name="Calibri", size=10)
    TITLE_FONT   = Font(name="Calibri", color="FFFFFF", bold=True, size=16)

    thin         = Side(style="thin",   color="B0B0B0")
    THIN_BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    total = len(TEST_CASES)

    # ══ SHEET 1: Summary ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells("A1:G1")
    ws["A1"].value     = "CrowdSense — Appium Mobile E2E Test Report"
    ws["A1"].font      = TITLE_FONT
    ws["A1"].fill      = TITLE_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 45

    ws.merge_cells("A2:G2")
    ws["A2"].value = (
        f"Device: Pixel 3a API 37  |  Platform: Android  |  "
        f"Automation: UiAutomator2  |  Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    ws["A2"].font      = Font(name="Calibri", color="FFFFFF", size=10)
    ws["A2"].fill      = PatternFill("solid", fgColor="2D5F8A")
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 10  # spacer

    # Info table
    info_hdrs = [("A4", "Property"), ("B4", "Value")]
    for cell_ref, hdr in info_hdrs:
        ws[cell_ref].value     = hdr
        ws[cell_ref].font      = WHITE_FONT
        ws[cell_ref].fill      = HEADER_FILL
        ws[cell_ref].alignment = CENTER
        ws[cell_ref].border    = THIN_BORDER

    info_rows = [
        ("Application Name",  "CrowdSense"),
        ("App Type",          "Flutter Android App"),
        ("Test Framework",    "Appium 2 + pytest (UiAutomator2)"),
        ("Test Environment",  "Pixel 3a API 37 AVD (Android Emulator)"),
        ("APK",               "frontend/build/app/outputs/flutter-apk/app-debug.apk"),
        ("Total Test Cases",  str(total)),
        ("Report Generated",  datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Prepared By",       "CrowdSense Automated Appium Suite"),
    ]
    for r_idx, (k, v) in enumerate(info_rows, start=5):
        ws[f"A{r_idx}"].value  = k
        ws[f"B{r_idx}"].value  = v
        ws[f"A{r_idx}"].font   = BOLD_FONT
        ws[f"B{r_idx}"].font   = NORMAL_FONT
        ws[f"A{r_idx}"].border = THIN_BORDER
        ws[f"B{r_idx}"].border = THIN_BORDER
        if r_idx % 2 == 0:
            ws[f"A{r_idx}"].fill = ALT_FILL
            ws[f"B{r_idx}"].fill = ALT_FILL

    # Stats table
    ws["D4"].value = "Status";     ws["D4"].font = WHITE_FONT; ws["D4"].fill = HEADER_FILL
    ws["E4"].value = "Count";      ws["E4"].font = WHITE_FONT; ws["E4"].fill = HEADER_FILL
    ws["F4"].value = "Percentage"; ws["F4"].font = WHITE_FONT; ws["F4"].fill = HEADER_FILL
    for cell in ["D4", "E4", "F4"]:
        ws[cell].alignment = CENTER; ws[cell].border = THIN_BORDER

    stats = [
        ("PASS",    0,     PASS_FILL,    Font(name="Calibri", color="276221", bold=True, size=10)),
        ("FAIL",    0,     FAIL_FILL,    Font(name="Calibri", color="9C0006", bold=True, size=10)),
        ("PENDING", total, PENDING_FILL, Font(name="Calibri", color="9C6500", bold=True, size=10)),
        ("TOTAL",   total, None,         BOLD_FONT),
    ]
    for r_idx, (label, count, fill, font) in enumerate(stats, start=5):
        pct = f"{(count / total * 100):.1f}%" if total > 0 else "0.0%"
        for col, val in [("D", label), ("E", count), ("F", pct)]:
            c = ws[f"{col}{r_idx}"]
            c.value = val; c.font = font; c.border = THIN_BORDER; c.alignment = CENTER
            if fill:
                c.fill = fill

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14

    # ══ SHEET 2: Test Results ════════════════════════════════════════════════
    ws2 = wb.create_sheet("Test Results")

    ws2.merge_cells("A1:I1")
    ws2["A1"].value     = "CrowdSense — Appium Mobile E2E Test Cases & Results"
    ws2["A1"].font      = TITLE_FONT
    ws2["A1"].fill      = TITLE_FILL
    ws2["A1"].alignment = CENTER
    ws2.row_dimensions[1].height = 38

    headers    = ["TC ID", "Test Module", "Test Case Description", "Category",
                  "Priority", "Status", "Actual Result / Notes", "Duration (s)", "Timestamp"]
    col_widths = [10, 30, 62, 22, 18, 12, 52, 14, 22]
    for col_idx, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws2.cell(row=2, column=col_idx, value=hdr)
        cell.font = WHITE_FONT; cell.fill = HEADER_FILL
        cell.border = THIN_BORDER; cell.alignment = CENTER
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[2].height = 30

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    for row_idx, (tc_id, module, description, category, priority) in enumerate(TEST_CASES, start=3):
        row_fill = ALT_FILL if row_idx % 2 == 0 else None
        values   = [tc_id, module, description, category, priority,
                    "PENDING", "— Not yet executed —", "", now_str]
        for col_idx, val in enumerate(values, start=1):
            c = ws2.cell(row=row_idx, column=col_idx, value=val)
            c.font = NORMAL_FONT; c.border = THIN_BORDER; c.alignment = LEFT
            if row_fill:
                c.fill = row_fill
        status_cell           = ws2.cell(row=row_idx, column=6)
        status_cell.fill      = PENDING_FILL
        status_cell.font      = Font(name="Calibri", color="9C6500", bold=True, size=10)
        status_cell.alignment = CENTER
        ws2.row_dimensions[row_idx].height = 22

    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:I{len(TEST_CASES) + 2}"

    # ══ SHEET 3: Test Scenarios ══════════════════════════════════════════════
    ws3 = wb.create_sheet("Test Scenarios")
    ws3.merge_cells("A1:E1")
    ws3["A1"].value     = "CrowdSense — Appium Mobile Test Scenario Coverage"
    ws3["A1"].font      = TITLE_FONT
    ws3["A1"].fill      = TITLE_FILL
    ws3["A1"].alignment = CENTER
    ws3.row_dimensions[1].height = 38

    scen_hdrs = ["Feature Area", "No. of Tests", "Coverage", "Notes", "Status"]
    for col_idx, h in enumerate(scen_hdrs, start=1):
        c = ws3.cell(row=2, column=col_idx, value=h)
        c.font = WHITE_FONT; c.fill = HEADER_FILL
        c.border = THIN_BORDER; c.alignment = CENTER
    ws3.row_dimensions[2].height = 28

    scenarios = [
        ("Authentication (Login, Register, Forgot Password)", 22, "100%",
         "All auth flows, validation, route guard, successful login", "Pending"),
        ("Home Screen & Navigation", 18, "100%",
         "Search bar, categories, trending, recent, bottom nav tabs", "Pending"),
        ("Search & Place Discovery", 7, "100%",
         "Search input, suggestions, recent searches, results list", "Pending"),
        ("Place Details", 9, "100%",
         "Name, location, crowd status, best time, forecast, about", "Pending"),
        ("Crowd Intelligence", 6, "100%",
         "Crowd status card, forecast, report form (low/med/high), note", "Pending"),
        ("Community Photos", 1, "100%",
         "Community photos page navigation and load", "Pending"),
        ("Profile", 11, "100%",
         "User stats, menu items, edit profile, sign out, my reports", "Pending"),
        ("Settings", 7, "100%",
         "Dark mode, notifications, privacy, terms, version, delete account", "Pending"),
        ("Travel Planner", 6, "100%",
         "Heading, from/to selectors, calculate button, crowd-aware text", "Pending"),
        ("Favorites", 6, "100%",
         "Favorites screen, empty state, add/remove favorites, icon toggle", "Pending"),
        ("Smoke / Critical Path", 10, "100%",
         "App launch, all screen access, E2E flows, rapid transitions", "Pending"),
    ]
    for r_idx, row in enumerate(scenarios, start=3):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, start=1):
            c = ws3.cell(row=r_idx, column=col_idx, value=val)
            c.font = NORMAL_FONT; c.border = THIN_BORDER; c.alignment = LEFT
            if fill:
                c.fill = fill
        ws3.row_dimensions[r_idx].height = 22

    ws3.column_dimensions["A"].width = 52
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 60
    ws3.column_dimensions["E"].width = 12

    wb.save(output_path)
    print(f"[OK] Appium XLSX report generated → {output_path}")
    return output_path


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    path = generate_xlsx_report(sys.argv[1] if len(sys.argv) > 1 else None)
    print(f"\nOpen the report at: {os.path.abspath(path)}")
