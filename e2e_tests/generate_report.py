"""
CrowdSense — Standalone XLSX Test Report Generator

Run this script AFTER running pytest to regenerate a pre-filled XLSX report
with placeholder "PENDING" results if no tests have been run yet, or to
create an initial report template for manual review.

Usage:
    python generate_report.py
"""
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://thirulogasundar.github.io/CrowdSense"

# ─── All 120 test cases defined statically for the template ──────────────────
TEST_CASES = [
    # Auth Tests
    ("TC001", "test_01_auth", "App loads and shows splash / login page", "Auth"),
    ("TC002", "test_01_auth", "Login page renders email, password and Sign In button", "Auth"),
    ("TC003", "test_01_auth", "Submitting empty login form shows validation error", "Auth"),
    ("TC004", "test_01_auth", "Invalid email format triggers validation", "Auth"),
    ("TC005", "test_01_auth", "Wrong credentials display error snackbar", "Auth"),
    ("TC006", "test_01_auth", "Clicking Forgot Password navigates to reset screen", "Auth"),
    ("TC007", "test_01_auth", "Forgot Password page shows email field and submit button", "Auth"),
    ("TC008", "test_01_auth", "Submitting empty email on forgot password shows validation", "Auth"),
    ("TC009", "test_01_auth", "Register page renders Full Name, Email, Password, Confirm", "Auth"),
    ("TC010", "test_01_auth", "Submitting empty register form keeps user on register page", "Auth"),
    ("TC011", "test_01_auth", "Mismatched confirm password shows error", "Auth"),
    ("TC012", "test_01_auth", "Invalid email format in register shows validation", "Auth"),
    ("TC013", "test_01_auth", "Unauthenticated user redirected to login from /home", "Auth"),
    ("TC014", "test_01_auth", "Sign In link on register page navigates to login", "Auth"),
    ("TC015", "test_01_auth", "Sign Up link on login page navigates to register", "Auth"),
    ("TC016", "test_01_auth", "Login page shows Welcome Back heading", "Auth"),
    ("TC017", "test_01_auth", "Register page shows Create Account heading", "Auth"),
    ("TC018", "test_01_auth", "Rapid auth page transitions don't crash app", "Auth"),
    # Home & Navigation Tests
    ("TC019", "test_02_home_navigation", "Home page loads or redirects properly", "Navigation"),
    ("TC020", "test_02_home_navigation", "Login has CrowdSense branding", "Navigation"),
    ("TC021", "test_02_home_navigation", "Home page contains greeting text", "Navigation"),
    ("TC022", "test_02_home_navigation", "Bottom navigation bar rendered", "Navigation"),
    ("TC023", "test_02_home_navigation", "Search bar visible on home page", "Navigation"),
    ("TC024", "test_02_home_navigation", "Trending Now section visible", "Navigation"),
    ("TC025", "test_02_home_navigation", "Popular Categories section visible", "Navigation"),
    ("TC026", "test_02_home_navigation", "Landmarks category chip present", "Navigation"),
    ("TC027", "test_02_home_navigation", "Restaurants category chip present", "Navigation"),
    ("TC028", "test_02_home_navigation", "Parks category chip present", "Navigation"),
    ("TC029", "test_02_home_navigation", "Shopping category chip present", "Navigation"),
    ("TC030", "test_02_home_navigation", "Entertainment category chip present", "Navigation"),
    ("TC031", "test_02_home_navigation", "Recent Searches section visible", "Navigation"),
    ("TC032", "test_02_home_navigation", "Search bar navigates to /search", "Navigation"),
    ("TC033", "test_02_home_navigation", "Navigating to /profile loads profile page", "Navigation"),
    ("TC034", "test_02_home_navigation", "Navigating to /planner loads Travel Planner", "Navigation"),
    ("TC035", "test_02_home_navigation", "Navigating to /favorites loads Favorites", "Navigation"),
    ("TC036", "test_02_home_navigation", "Navigating to /settings loads Settings", "Navigation"),
    ("TC037", "test_02_home_navigation", "Settings Dark Mode toggle present", "Navigation"),
    ("TC038", "test_02_home_navigation", "Settings Notifications toggle present", "Navigation"),
    ("TC039", "test_02_home_navigation", "Settings App Version present", "Navigation"),
    ("TC040", "test_02_home_navigation", "Delete Account option in Settings", "Navigation"),
    # Search & Place Tests
    ("TC041", "test_03_search_place_crowd", "Search page loads with input field", "Search"),
    ("TC042", "test_03_search_place_crowd", "Initial search shows suggestion chips", "Search"),
    ("TC043", "test_03_search_place_crowd", "Search results page loads", "Search"),
    ("TC044", "test_03_search_place_crowd", "Search screen shows Recent Searches", "Search"),
    ("TC045", "test_03_search_place_crowd", "Place details page loads with place ID", "Place Details"),
    ("TC046", "test_03_search_place_crowd", "Place details shows loader or content", "Place Details"),
    ("TC047", "test_03_search_place_crowd", "Crowd Status card shown on place details", "Crowd"),
    ("TC048", "test_03_search_place_crowd", "Best Time to Visit shown on place details", "Crowd"),
    ("TC049", "test_03_search_place_crowd", "Crowd Forecast chart visible", "Crowd"),
    ("TC050", "test_03_search_place_crowd", "Report Live Crowd Level button present", "Crowd"),
    ("TC051", "test_03_search_place_crowd", "Favorite heart icon in place details", "Place Details"),
    ("TC052", "test_03_search_place_crowd", "Share icon in place details", "Place Details"),
    ("TC053", "test_03_search_place_crowd", "About section in place details", "Place Details"),
    ("TC054", "test_03_search_place_crowd", "Community Photos page loads", "Community"),
    ("TC055", "test_03_search_place_crowd", "Crowd report has Not busy / Low option", "Crowd"),
    ("TC056", "test_03_search_place_crowd", "Crowd report has A bit busy / Moderate option", "Crowd"),
    ("TC057", "test_03_search_place_crowd", "Crowd report has Very crowded / High option", "Crowd"),
    ("TC058", "test_03_search_place_crowd", "Crowd report has optional note field", "Crowd"),
    ("TC059", "test_03_search_place_crowd", "Crowd report Submit button present", "Crowd"),
    ("TC060", "test_03_search_place_crowd", "Place details shows location info", "Place Details"),
    ("TC061", "test_03_search_place_crowd", "Search loading state handled", "Search"),
    ("TC062", "test_03_search_place_crowd", "Search shows empty results state", "Search"),
    ("TC063", "test_03_search_place_crowd", "Place name shown on place details", "Place Details"),
    ("TC064", "test_03_search_place_crowd", "Place category badge displayed", "Place Details"),
    ("TC065", "test_03_search_place_crowd", "Place image header in SliverAppBar", "Place Details"),
    # Profile, Planner & Settings Tests
    ("TC066", "test_04_profile_planner_settings", "Profile page renders", "Profile"),
    ("TC067", "test_04_profile_planner_settings", "Profile shows Photos stat", "Profile"),
    ("TC068", "test_04_profile_planner_settings", "Profile shows Reports stat", "Profile"),
    ("TC069", "test_04_profile_planner_settings", "Profile shows Saved stat", "Profile"),
    ("TC070", "test_04_profile_planner_settings", "Edit Profile menu item present", "Profile"),
    ("TC071", "test_04_profile_planner_settings", "Notifications menu item in profile", "Profile"),
    ("TC072", "test_04_profile_planner_settings", "My Reports menu item in profile", "Profile"),
    ("TC073", "test_04_profile_planner_settings", "Support and FAQ menu in profile", "Profile"),
    ("TC074", "test_04_profile_planner_settings", "Sign Out button in profile", "Profile"),
    ("TC075", "test_04_profile_planner_settings", "Settings icon in profile app bar", "Profile"),
    ("TC076", "test_04_profile_planner_settings", "My Reports page loads", "Profile"),
    ("TC077", "test_04_profile_planner_settings", "Favorites page empty state shown", "Favorites"),
    ("TC078", "test_04_profile_planner_settings", "Travel Planner renders heading", "Travel Planner"),
    ("TC079", "test_04_profile_planner_settings", "Travel Planner has From selector", "Travel Planner"),
    ("TC080", "test_04_profile_planner_settings", "Travel Planner has To selector", "Travel Planner"),
    ("TC081", "test_04_profile_planner_settings", "Travel Planner Calculate Best Plan button", "Travel Planner"),
    ("TC082", "test_04_profile_planner_settings", "Travel Planner crowd-aware subtitle", "Travel Planner"),
    ("TC083", "test_04_profile_planner_settings", "Settings Privacy Policy option", "Settings"),
    ("TC084", "test_04_profile_planner_settings", "Settings Terms of Service option", "Settings"),
    ("TC085", "test_04_profile_planner_settings", "Settings APPEARANCE section header", "Settings"),
    ("TC086", "test_04_profile_planner_settings", "Settings ABOUT section header", "Settings"),
    ("TC087", "test_04_profile_planner_settings", "Settings ACCOUNT section header", "Settings"),
    ("TC088", "test_04_profile_planner_settings", "App version shows 1.0.0", "Settings"),
    ("TC089", "test_04_profile_planner_settings", "Profile photo upload area visible", "Profile"),
    ("TC090", "test_04_profile_planner_settings", "My Reports shows report list or empty", "Profile"),
    # UI/UX & Edge Cases
    ("TC091", "test_05_ui_performance_edge", "Browser tab title is crowdsense", "UI/UX"),
    ("TC092", "test_05_ui_performance_edge", "App renders within 10 seconds", "Performance"),
    ("TC093", "test_05_ui_performance_edge", "No critical JS errors on login page", "UI/UX"),
    ("TC094", "test_05_ui_performance_edge", "App renders on mobile viewport 375x812", "UI/UX"),
    ("TC095", "test_05_ui_performance_edge", "App renders on tablet viewport 768x1024", "UI/UX"),
    ("TC096", "test_05_ui_performance_edge", "App renders on desktop viewport 1920x1080", "UI/UX"),
    ("TC097", "test_05_ui_performance_edge", "Browser back/forward navigation works", "UI/UX"),
    ("TC098", "test_05_ui_performance_edge", "Page refresh doesn't crash app", "UI/UX"),
    ("TC099", "test_05_ui_performance_edge", "Unknown route handled gracefully", "Edge Cases"),
    ("TC100", "test_05_ui_performance_edge", "Multiple rapid URL changes don't crash", "Edge Cases"),
    ("TC101", "test_05_ui_performance_edge", "Scroll on login page works", "UI/UX"),
    ("TC102", "test_05_ui_performance_edge", "Scroll on register page works", "UI/UX"),
    ("TC103", "test_05_ui_performance_edge", "App favicon loads with 200 status", "Performance"),
    ("TC104", "test_05_ui_performance_edge", "manifest.json accessible from base URL", "Performance"),
    ("TC105", "test_05_ui_performance_edge", "flutter_bootstrap.js accessible", "Performance"),
    ("TC106", "test_05_ui_performance_edge", "Login page accepts keyboard tab input", "Accessibility"),
    ("TC107", "test_05_ui_performance_edge", "Invalid place ID doesn't crash app", "Edge Cases"),
    ("TC108", "test_05_ui_performance_edge", "Splash screen auto-redirects within 5s", "Navigation"),
    ("TC109", "test_05_ui_performance_edge", "All auth routes accessible without state", "Auth"),
    ("TC110", "test_05_ui_performance_edge", "Protected routes redirect unauthenticated", "Auth"),
    ("TC111", "test_05_ui_performance_edge", "Register accessible from login", "Auth"),
    ("TC112", "test_05_ui_performance_edge", "Page layout stable, no horizontal overflow", "UI/UX"),
    ("TC113", "test_05_ui_performance_edge", "flutter.js accessible from base URL", "Performance"),
    ("TC114", "test_05_ui_performance_edge", "Forgot password shows Send Reset Link text", "Auth"),
    ("TC115", "test_05_ui_performance_edge", "App page source is valid HTML", "UI/UX"),
    # Smoke Tests
    ("TC116", "test_06_smoke", "Smoke – /login route loads with Sign In", "Smoke"),
    ("TC117", "test_06_smoke", "Smoke – /register route loads with Create Account", "Smoke"),
    ("TC118", "test_06_smoke", "Smoke – /forgot-password loads with Reset", "Smoke"),
    ("TC119", "test_06_smoke", "Smoke – /settings loads with Settings", "Smoke"),
    ("TC120", "test_06_smoke", "Smoke – App base URL loads Flutter app shell", "Smoke"),
]


def generate_xlsx_report(output_path=None):
    """Generate an XLSX test report template with all 120 test cases."""
    if output_path is None:
        ts = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
        output_path = f"E2E_Test_Report_CrowdSense_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # ─── Colour definitions ────────────────────────────────────────────────
    PASS_FILL    = PatternFill("solid", fgColor="C6EFCE")
    FAIL_FILL    = PatternFill("solid", fgColor="FFC7CE")
    PENDING_FILL = PatternFill("solid", fgColor="FFEB9C")
    HEADER_FILL  = PatternFill("solid", fgColor="2D5F8A")
    TITLE_FILL   = PatternFill("solid", fgColor="1A3A5C")
    ALT_FILL     = PatternFill("solid", fgColor="EBF2FA")
    SUBHDR_FILL  = PatternFill("solid", fgColor="4472C4")

    WHITE_FONT   = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
    BOLD_FONT    = Font(name="Calibri", bold=True, size=11)
    NORMAL_FONT  = Font(name="Calibri", size=10)
    TITLE_FONT   = Font(name="Calibri", color="FFFFFF", bold=True, size=16)

    thin  = Side(style="thin",   color="B0B0B0")
    thick = Side(style="medium", color="2D5F8A")
    THIN_BORDER  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    THICK_BORDER = Border(left=thick, right=thick, top=thick, bottom=thick)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    total   = len(TEST_CASES)
    passed  = sum(1 for _ in TEST_CASES)  # All PENDING for template
    pending = total

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: Summary
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"].value     = "CrowdSense -- End-to-End Test Execution Report"
    ws["A1"].font      = TITLE_FONT
    ws["A1"].fill      = TITLE_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 45

    # Subtitle
    ws.merge_cells("A2:G2")
    ws["A2"].value     = f"Application: CrowdSense  |  URL: {BASE_URL}  |  Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font      = Font(name="Calibri", color="FFFFFF", size=10)
    ws["A2"].fill      = PatternFill("solid", fgColor="2D5F8A")
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 10  # spacer

    # Execution Info table
    info_header_cells = ["A4", "B4"]
    for cell in info_header_cells:
        ws[cell].font  = WHITE_FONT
        ws[cell].fill  = HEADER_FILL
        ws[cell].alignment = CENTER
        ws[cell].border = THIN_BORDER
    ws["A4"].value = "Property"
    ws["B4"].value = "Value"

    info_rows = [
        ("Application Name",  "CrowdSense"),
        ("App Type",          "Flutter Web App"),
        ("Test Framework",    "Selenium 4 + pytest"),
        ("Test Environment",  "GitHub Pages (Production)"),
        ("Base URL",          BASE_URL),
        ("Total Test Cases",  str(total)),

        ("Report Generated",  datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Prepared By",       "Automated Test Suite"),
    ]
    for r_idx, (k, v) in enumerate(info_rows, start=5):
        ws[f"A{r_idx}"].value  = k
        ws[f"B{r_idx}"].value  = v
        ws[f"A{r_idx}"].font   = BOLD_FONT
        ws[f"B{r_idx}"].font   = NORMAL_FONT
        ws[f"A{r_idx}"].border = THIN_BORDER
        ws[f"B{r_idx}"].border = THIN_BORDER
        if r_idx % 2 == 0:
            ws[f"A{r_idx}"].fill = ALT_FILL
            ws[f"B{r_idx}"].fill = ALT_FILL

    ws.row_dimensions[13].height = 10  # spacer

    # Statistics table
    ws["D4"].value = "Status";    ws["D4"].font = WHITE_FONT; ws["D4"].fill = HEADER_FILL
    ws["E4"].value = "Count";     ws["E4"].font = WHITE_FONT; ws["E4"].fill = HEADER_FILL
    ws["F4"].value = "Percentage";ws["F4"].font = WHITE_FONT; ws["F4"].fill = HEADER_FILL
    for cell in ["D4", "E4", "F4"]:
        ws[cell].alignment = CENTER; ws[cell].border = THIN_BORDER

    stats = [
        ("PASS",    0,      PASS_FILL,    Font(name="Calibri", color="276221", bold=True, size=10)),
        ("FAIL",    0,      FAIL_FILL,    Font(name="Calibri", color="9C0006", bold=True, size=10)),
        ("PENDING", total,  PENDING_FILL, Font(name="Calibri", color="9C6500", bold=True, size=10)),
        ("TOTAL",     total,  None,         BOLD_FONT),
    ]
    for r_idx, (label, count, fill, font) in enumerate(stats, start=5):
        pct = f"{(count/total*100):.1f}%" if total > 0 else "0.0%"
        for col, val in [("D", label), ("E", count), ("F", pct)]:
            c = ws[f"{col}{r_idx}"]
            c.value = val; c.font = font; c.border = THIN_BORDER; c.alignment = CENTER
            if fill:
                c.fill = fill

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2: Test Results
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Test Results")

    # Title
    ws2.merge_cells("A1:I1")
    ws["A1"].value     = "CrowdSense -- E2E Test Cases & Results"
    ws2["A1"].font      = TITLE_FONT
    ws2["A1"].fill      = TITLE_FILL
    ws2["A1"].alignment = CENTER
    ws2.row_dimensions[1].height = 38

    headers = ["TC ID", "Test Module", "Test Case Description", "Category",
               "Priority", "Status", "Actual Result / Notes",
               "Execution Time (s)", "Timestamp"]
    col_widths = [10, 28, 58, 18, 12, 12, 50, 18, 22]

    for col_idx, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws2.cell(row=2, column=col_idx, value=hdr)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.border    = THIN_BORDER
        cell.alignment = CENTER
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[2].height = 30

    # Determine priority based on module
    def priority(module):
        if "auth" in module or "smoke" in module:
            return "P1 – Critical"
        if "home" in module or "search" in module:
            return "P2 – High"
        if "profile" in module or "planner" in module:
            return "P3 – Medium"
        return "P4 – Low"

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    for row_idx, (tc_id, module, description, category) in enumerate(TEST_CASES, start=3):
        row_fill = ALT_FILL if row_idx % 2 == 0 else None
        values = [tc_id, module, description, category,
                  priority(module), "PENDING", "— Not yet executed —", "", now_str]
        for col_idx, val in enumerate(values, start=1):
            c = ws2.cell(row=row_idx, column=col_idx, value=val)
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = LEFT
            if row_fill:
                c.fill = row_fill
        # Status cell colour
        status_cell = ws2.cell(row=row_idx, column=6)
        status_cell.fill      = PENDING_FILL
        status_cell.font      = Font(name="Calibri", color="9C6500", bold=True, size=10)
        status_cell.alignment = CENTER
        ws2.row_dimensions[row_idx].height = 22

    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:I{len(TEST_CASES) + 2}"

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3: Test Scenarios (by feature area)
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Test Scenarios")
    ws3.merge_cells("A1:E1")
    ws3["A1"].value     = "CrowdSense — Test Scenario Coverage"
    ws3["A1"].font      = TITLE_FONT
    ws3["A1"].fill      = TITLE_FILL
    ws3["A1"].alignment = CENTER
    ws3.row_dimensions[1].height = 38

    scenario_headers = ["Feature Area", "No. of Test Cases", "Coverage", "Notes", "Status"]
    for col_idx, h in enumerate(scenario_headers, start=1):
        c = ws3.cell(row=2, column=col_idx, value=h)
        c.font = WHITE_FONT; c.fill = HEADER_FILL
        c.border = THIN_BORDER; c.alignment = CENTER
    ws3.row_dimensions[2].height = 28

    scenarios = [
        ("Authentication (Login, Register, Forgot Password)", 18, "100%",
         "All auth flows including validation and route guard", "Pending"),
        ("Home Screen & Navigation", 22, "100%",
         "Bottom nav, categories, trending, recent searches", "Pending"),
        ("Search & Place Discovery", 10, "100%",
         "Search input, results list, suggestions", "Pending"),
        ("Place Details", 9, "100%",
         "Full details, image header, location, category", "Pending"),
        ("Crowd Intelligence", 7, "100%",
         "Live crowd status, forecast chart, crowd report form", "Pending"),
        ("Community Photos", 1, "100%",
         "Community photos page loads", "Pending"),
        ("Profile", 11, "100%",
         "Stats, menus, edit profile, sign out", "Pending"),
        ("Favorites", 2, "100%",
         "Favorites page and empty state", "Pending"),
        ("Travel Planner", 5, "100%",
         "Source/destination selectors, map, calculate button", "Pending"),
        ("Settings", 10, "100%",
         "Dark mode, notifications, privacy, terms, version", "Pending"),
        ("UI/UX & Responsiveness", 8, "100%",
         "Mobile, tablet, desktop viewports, scroll, layout", "Pending"),
        ("Performance", 5, "100%",
         "Load time, static asset checks", "Pending"),
        ("Edge Cases", 4, "100%",
         "Unknown route, rapid navigation, invalid place ID", "Pending"),
        ("Smoke Tests", 5, "100%",
         "Quick sanity for all critical routes", "Pending"),
        ("Accessibility", 3, "100%",
         "Keyboard navigation, tab order", "Pending"),
    ]
    for r_idx, row in enumerate(scenarios, start=3):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for col_idx, val in enumerate(row, start=1):
            c = ws3.cell(row=r_idx, column=col_idx, value=val)
            c.font = NORMAL_FONT; c.border = THIN_BORDER; c.alignment = LEFT
            if fill:
                c.fill = fill
        ws3.row_dimensions[r_idx].height = 22

    ws3.column_dimensions["A"].width = 52
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 55
    ws3.column_dimensions["E"].width = 16

    wb.save(output_path)
    print(f"[OK] XLSX report generated -> {output_path}")
    return output_path


if __name__ == "__main__":
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    path = generate_xlsx_report()
    print(f"\nOpen the report at: {path}")
