"""
CrowdSense E2E Test Configuration (conftest.py)
================================================
- Selenium Chrome driver fixture (module-scoped for speed)
- Pytest hooks to auto-generate XLSX report matching the PancreaScan reference format
  Sheets: Summary | Passed Tests | Failed Tests | Execution Log | Test Details
"""
import pytest
import time
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
try:
    from webdriver_manager.chrome import ChromeDriverManager
    USE_WDM = True
except ImportError:
    USE_WDM = False

# ── App URL ──────────────────────────────────────────────────────────────────
BASE_URL = "https://thirulogasundar.github.io/CrowdSense"

# ── Reference palette (matched from PancreaScan xlsx) ────────────────────────
HDR_FILL    = PatternFill("solid", fgColor="1F3864")   # dark navy header
PASS_FILL   = PatternFill("solid", fgColor="C6EFCE")   # light green
FAIL_FILL   = PatternFill("solid", fgColor="FFC7CE")   # light red
WHITE_FONT  = Font(name="Calibri", color="FFFFFF", bold=True)
BOLD_FONT   = Font(name="Calibri", bold=True)
NORMAL_FONT = Font(name="Calibri")
thin = Side(style="thin", color="B0B0B0")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Global result stores ─────────────────────────────────────────────────────
_suite_start = None
_results = []    # list of dicts: {no, category, name, status, duration, error, ts}


def pytest_sessionstart(session):
    global _suite_start
    _suite_start = datetime.datetime.utcnow()


def pytest_runtest_logreport(report):
    global _results
    if report.when == "call":
        status  = "PASSED" if report.passed else ("FAILED" if report.failed else "SKIPPED")
        error   = str(report.longrepr)[:600] if report.failed else "None — test passed successfully."
        ts      = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _results.append({
            "no":       len(_results) + 1,
            "category": _get_category(report.nodeid),
            "name":     report.nodeid.split("::")[-1],
            "status":   status,
            "duration": round(report.duration, 2),
            "error":    error,
            "ts":       ts,
        })


def _get_category(nodeid):
    mapping = {
        "test_01_auth":                    "Authentication",
        "test_02_home_navigation":         "Home & Navigation",
        "test_03_search_place_crowd":      "Search & Crowd",
        "test_04_profile_planner_settings":"Profile & Settings",
        "test_05_ui_performance_edge":     "UI/UX & Performance",
        "test_06_smoke":                   "Smoke Tests",
    }
    for key, cat in mapping.items():
        if key in nodeid:
            return cat
    return "General"


def pytest_sessionfinish(session, exitstatus):
    _generate_xlsx()


# ── XLSX writer ───────────────────────────────────────────────────────────────
def _generate_xlsx():
    if not _results:
        return

    suite_end  = datetime.datetime.utcnow()
    suite_start_str = _suite_start.strftime("%Y-%m-%dT%H:%M:%S.%f") + "Z" if _suite_start else ""
    suite_end_str   = suite_end.strftime("%Y-%m-%dT%H:%M:%S.%f") + "Z"

    passed_rows  = [r for r in _results if r["status"] == "PASSED"]
    failed_rows  = [r for r in _results if r["status"] == "FAILED"]
    total        = len(_results)
    passed_count = len(passed_rows)
    failed_count = len(failed_rows)
    pass_rate    = round(passed_count / total * 100, 2) if total else 0
    duration_sec = sum(r["duration"] for r in _results)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_header(ws_sum, ["Test Suite", "Total Tests", "Passed", "Failed",
                            "Pass Rate %", "Duration (sec)", "Start Time", "End Time"])
    row = [
        "CrowdSense Web App — Full E2E Workflow",
        total, passed_count, failed_count,
        pass_rate, round(duration_sec, 2),
        suite_start_str, suite_end_str,
    ]
    for col_idx, val in enumerate(row, start=1):
        c = ws_sum.cell(row=2, column=col_idx, value=val)
        c.font   = NORMAL_FONT
        c.border = THIN_BORDER
        c.alignment = LEFT
    ws_sum.column_dimensions["A"].width = 50
    for col in ["B","C","D","E","F"]:
        ws_sum.column_dimensions[col].width = 16
    ws_sum.column_dimensions["G"].width = 30
    ws_sum.column_dimensions["H"].width = 30

    # ── Sheet 2: Passed Tests ─────────────────────────────────────────────────
    ws_pass = wb.create_sheet("Passed Tests")
    _write_header(ws_pass, ["No.", "Category", "Test Name", "Time (sec)", "Status"])
    for r_idx, r in enumerate(passed_rows, start=2):
        vals = [r["no"], r["category"], r["name"], r["duration"], r["status"]]
        for c_idx, val in enumerate(vals, start=1):
            c = ws_pass.cell(row=r_idx, column=c_idx, value=val)
            c.font   = NORMAL_FONT
            c.fill   = PASS_FILL
            c.border = THIN_BORDER
            c.alignment = LEFT
    ws_pass.column_dimensions["A"].width = 6
    ws_pass.column_dimensions["B"].width = 24
    ws_pass.column_dimensions["C"].width = 65
    ws_pass.column_dimensions["D"].width = 14
    ws_pass.column_dimensions["E"].width = 12

    # ── Sheet 3: Failed Tests ─────────────────────────────────────────────────
    ws_fail = wb.create_sheet("Failed Tests")
    _write_header(ws_fail, ["No.", "Category", "Test Name", "Error", "Status", "Timestamp"])
    for r_idx, r in enumerate(failed_rows, start=2):
        vals = [r["no"], r["category"], r["name"], r["error"], r["status"], r["ts"]]
        for c_idx, val in enumerate(vals, start=1):
            c = ws_fail.cell(row=r_idx, column=c_idx, value=val)
            c.font   = NORMAL_FONT
            c.fill   = FAIL_FILL
            c.border = THIN_BORDER
            c.alignment = LEFT
    ws_fail.column_dimensions["A"].width = 6
    ws_fail.column_dimensions["B"].width = 24
    ws_fail.column_dimensions["C"].width = 50
    ws_fail.column_dimensions["D"].width = 80
    ws_fail.column_dimensions["E"].width = 12
    ws_fail.column_dimensions["F"].width = 22

    # ── Sheet 4: Execution Log ────────────────────────────────────────────────
    ws_log = wb.create_sheet("Execution Log")
    _write_header(ws_log, ["Timestamp", "Level", "Message"])
    for r_idx, r in enumerate(_results, start=2):
        lvl = "INFO" if r["status"] == "PASSED" else "ERROR"
        msg = f"[{r['category']}] {r['name']} -> {r['status']} in {r['duration']}s"
        for c_idx, val in enumerate([r["ts"], lvl, msg], start=1):
            c = ws_log.cell(row=r_idx, column=c_idx, value=val)
            c.font   = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = LEFT
    ws_log.column_dimensions["A"].width = 22
    ws_log.column_dimensions["B"].width = 10
    ws_log.column_dimensions["C"].width = 90

    # ── Sheet 5: Test Details ─────────────────────────────────────────────────
    ws_det = wb.create_sheet("Test Details")
    _write_header(ws_det, ["No.", "Category", "Test Name", "Status", "Error Details"])
    for r_idx, r in enumerate(_results, start=2):
        vals = [r["no"], r["category"], r["name"], r["status"], r["error"]]
        fill = PASS_FILL if r["status"] == "PASSED" else FAIL_FILL
        for c_idx, val in enumerate(vals, start=1):
            c = ws_det.cell(row=r_idx, column=c_idx, value=val)
            c.font   = NORMAL_FONT
            c.fill   = fill
            c.border = THIN_BORDER
            c.alignment = LEFT
    ws_det.column_dimensions["A"].width = 6
    ws_det.column_dimensions["B"].width = 24
    ws_det.column_dimensions["C"].width = 55
    ws_det.column_dimensions["D"].width = 12
    ws_det.column_dimensions["E"].width = 80

    ts = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    fname = f"E2E_Test_Report_CrowdSense_{ts}.xlsx"
    wb.save(fname)
    print(f"\n[REPORT] XLSX saved -> {fname}")


def _write_header(ws, columns):
    for col_idx, hdr in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=hdr)
        c.font      = WHITE_FONT
        c.fill      = HDR_FILL
        c.border    = THIN_BORDER
        c.alignment = CENTER
    ws.row_dimensions[1].height = 25


# ── Selenium fixture ──────────────────────────────────────────────────────────
@pytest.fixture(scope="module")
def driver():
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-notifications")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--log-level=3")
    # opts.add_argument("--headless=new")   # uncomment for headless mode
    if USE_WDM:
        service = Service(ChromeDriverManager().install())
    else:
        service = Service()
    drv = webdriver.Chrome(service=service, options=opts)
    drv.implicitly_wait(10)
    yield drv
    drv.quit()


@pytest.fixture
def base_url():
    return BASE_URL
