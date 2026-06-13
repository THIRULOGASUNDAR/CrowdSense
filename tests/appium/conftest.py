import os
import sys
import time
from datetime import datetime
import pytest
from appium import webdriver as appium_webdriver
from appium.options.android import UiAutomator2Options

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    pass

APPIUM_SERVER = "http://127.0.0.1:4723"
DEVICE_NAME = "10BD7N0XY30009F"

_results = []

def pytest_runtest_makereport(item, call):
    if call.when == "call":
        _results.append({
            "name": item.name,
            "status": call.excinfo is None and "PASSED" or "FAILED",
            "duration": round(call.duration, 2),
            "error": str(call.excinfo.value) if call.excinfo else ""
        })

def pytest_sessionfinish(session, exitstatus):
    report_dir = os.path.join(os.path.dirname(__file__), "reports")
    os.makedirs(report_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    excel_path = os.path.join(report_dir, f"CrowdSense_Appium_Report_{ts}.xlsx")
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Appium E2E Results"
        
        headers = ["Test Name", "Status", "Duration (s)", "Error Details"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            
        for row, res in enumerate(_results, 2):
            ws.cell(row=row, column=1, value=res["name"])
            
            c_status = ws.cell(row=row, column=2, value=res["status"])
            c_status.fill = PatternFill("solid", fgColor="C6EFCE" if res["status"] == "PASSED" else "FFC7CE")
            
            ws.cell(row=row, column=3, value=res["duration"])
            ws.cell(row=row, column=4, value=res["error"])
            
        wb.save(excel_path)
    except Exception as e:
        print(f"Failed to generate Excel report: {e}")

@pytest.fixture(scope="module")
def driver():
    options = UiAutomator2Options()
    options.platform_name = "Android"
    options.automation_name = "UiAutomator2"
    options.device_name = DEVICE_NAME
    options.app_package = "com.example.crowdsense"
    options.app_activity = "com.example.crowdsense.MainActivity"
    
    apk_path = os.path.join(os.path.dirname(__file__), "..", "build", "app", "outputs", "flutter-apk", "app-debug.apk")
    if os.path.exists(apk_path):
        options.app = os.path.abspath(apk_path)
        
    options.new_command_timeout = 180
    options.auto_grant_permissions = True
    
    drv = appium_webdriver.Remote(command_executor=APPIUM_SERVER, options=options)
    drv.implicitly_wait(0)
    yield drv
    drv.quit()
