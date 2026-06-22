"""
generate_report.py — Standalone Excel Report Generator
=======================================================
CrowdSense Selenium E2E Test Suite

Use this script to generate a pre-filled Excel template report
BEFORE running tests (for planning) or AFTER running tests
(combined with conftest auto-generation, for a static template).

Usage:
    python selenium_tests/generate_report.py
    python selenium_tests/generate_report.py --output MyReport.xlsx

Auto-generation:
    The conftest.py automatically generates a live report at the end
    of every `pytest` run. This script creates a static template.
"""
from __future__ import annotations

import argparse
import datetime
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://thirulogasundar.github.io/CrowdSense"

# ─── Colour Palette ───────────────────────────────────────────────────────────
_NAVY      = "1F3864"
_BLUE      = "2E75B6"
_PASS_BG   = "C6EFCE"
_FAIL_BG   = "FFC7CE"
_SKIP_BG   = "FFEB9C"
_PENDING   = "FFEB9C"
_ALT_BG    = "EBF2FA"
_WHITE     = "FFFFFF"


def _fill(c):
    return PatternFill("solid", fgColor=c)


def _font(color=_WHITE, bold=False, size=11, name="Calibri"):
    return Font(name=name, color=color, bold=bold, size=size)


def _border():
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)


CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

HDR_FILL   = _fill(_NAVY)
PASS_FILL  = _fill(_PASS_BG)
FAIL_FILL  = _fill(_FAIL_BG)
SKIP_FILL  = _fill(_SKIP_BG)
PEND_FILL  = _fill(_PENDING)
ALT_FILL   = _fill(_ALT_BG)
THIN_BDR   = _border()
WHITE_HDR  = _font(bold=True, size=11)
BOLD       = _font(color="000000", bold=True, size=10)
NORMAL     = _font(color="000000", size=10)

# ─── All 340 Test Cases ───────────────────────────────────────────────────────
TEST_CASES: list[tuple[str, str, str, str, str]] = [
    # (TC_ID, Module, Description, Category, Priority)
    # ── Authentication (TC001–TC025) ──────────────────────────────────────────
    ("TC001", "test_01_auth", "App base URL loads and Flutter canvas is rendered", "Authentication", "P1"),
    ("TC002", "test_01_auth", "Browser tab title contains 'crowdsense'", "Authentication", "P1"),
    ("TC003", "test_01_auth", "Splash screen auto-redirects to login within 6 seconds", "Authentication", "P1"),
    ("TC004", "test_01_auth", "Root URL serves a valid HTML document", "Authentication", "P1"),
    ("TC005", "test_01_auth", "/login route renders the Flutter app", "Authentication", "P1"),
    ("TC006", "test_01_auth", "Login page shows an email input field", "Authentication", "P1"),
    ("TC007", "test_01_auth", "Login page shows a password input field", "Authentication", "P1"),
    ("TC008", "test_01_auth", "Login page shows a Sign In button", "Authentication", "P1"),
    ("TC009", "test_01_auth", "Login page shows 'Welcome Back' heading", "Authentication", "P2"),
    ("TC010", "test_01_auth", "Login page has a Forgot Password link", "Authentication", "P2"),
    ("TC011", "test_01_auth", "Login page has a Sign Up link", "Authentication", "P2"),
    ("TC012", "test_01_auth", "Login page shows CrowdSense branding", "Authentication", "P2"),
    ("TC013", "test_01_auth", "Invalid credentials do not navigate away from login", "Authentication", "P2"),
    ("TC014", "test_01_auth", "/register route renders Flutter app", "Authentication", "P1"),
    ("TC015", "test_01_auth", "Register page shows Full Name input field", "Authentication", "P1"),
    ("TC016", "test_01_auth", "Register page shows Email input field", "Authentication", "P1"),
    ("TC017", "test_01_auth", "Register page shows Password input field", "Authentication", "P1"),
    ("TC018", "test_01_auth", "Register page shows Confirm Password field", "Authentication", "P1"),
    ("TC019", "test_01_auth", "Register page shows Create Account heading", "Authentication", "P2"),
    ("TC020", "test_01_auth", "Register page has a Sign In link", "Authentication", "P2"),
    ("TC021", "test_01_auth", "/forgot-password route renders Flutter app", "Authentication", "P1"),
    ("TC022", "test_01_auth", "Forgot Password page shows email input field", "Authentication", "P1"),
    ("TC023", "test_01_auth", "Forgot Password page shows Send Reset Link button", "Authentication", "P1"),
    ("TC024", "test_01_auth", "Unauthenticated /home is redirected to login", "Authentication", "P1"),
    ("TC025", "test_01_auth", "Unauthenticated /profile is redirected to login", "Authentication", "P1"),
    # ── Home (TC026–TC045) ────────────────────────────────────────────────────
    ("TC026", "test_02_home", "/home loads home page or redirects to login", "Home & Navigation", "P1"),
    ("TC027", "test_02_home", "Flutter canvas is rendered on home route", "Home & Navigation", "P1"),
    ("TC028", "test_02_home", "Unauthenticated user at /home sent to login", "Home & Navigation", "P1"),
    ("TC029", "test_02_home", "Home page search bar is present", "Home & Navigation", "P2"),
    ("TC030", "test_02_home", "Home page shows Trending Now section", "Home & Navigation", "P2"),
    ("TC031", "test_02_home", "Home page shows Popular Categories section", "Home & Navigation", "P2"),
    ("TC032", "test_02_home", "Home page shows Landmarks category chip", "Home & Navigation", "P2"),
    ("TC033", "test_02_home", "Home page shows Restaurants category chip", "Home & Navigation", "P2"),
    ("TC034", "test_02_home", "Home page shows Parks category chip", "Home & Navigation", "P2"),
    ("TC035", "test_02_home", "Home page shows Shopping category chip", "Home & Navigation", "P2"),
    ("TC036", "test_02_home", "Home page shows Entertainment category chip", "Home & Navigation", "P2"),
    ("TC037", "test_02_home", "Home page shows Recent Searches section", "Home & Navigation", "P2"),
    ("TC038", "test_02_home", "Browser title is 'crowdsense' on home route", "Home & Navigation", "P2"),
    ("TC039", "test_02_home", "Home page has a bottom navigation bar", "Home & Navigation", "P2"),
    ("TC040", "test_02_home", "Home page shows a greeting text", "Home & Navigation", "P3"),
    ("TC041", "test_02_home", "Home page does not show a 500 error", "Home & Navigation", "P1"),
    ("TC042", "test_02_home", "Home page does not show a 404 error", "Home & Navigation", "P1"),
    ("TC043", "test_02_home", "Home page source has substantial content", "Home & Navigation", "P2"),
    ("TC044", "test_02_home", "Scrolling on home page does not crash", "Home & Navigation", "P2"),
    ("TC045", "test_02_home", "Refreshing home route does not crash", "Home & Navigation", "P2"),
    # ── Search (TC046–TC065) ──────────────────────────────────────────────────
    ("TC046", "test_03_search", "/search route renders or redirects gracefully", "Search & Discovery", "P1"),
    ("TC047", "test_03_search", "Flutter rendered on /search route", "Search & Discovery", "P1"),
    ("TC048", "test_03_search", "Search page has a search input bar", "Search & Discovery", "P1"),
    ("TC049", "test_03_search", "Search page shows Coffee Shop chip", "Search & Discovery", "P2"),
    ("TC050", "test_03_search", "Search page shows Library chip", "Search & Discovery", "P2"),
    ("TC051", "test_03_search", "Search page shows Museum chip", "Search & Discovery", "P2"),
    ("TC052", "test_03_search", "Search page shows Mall chip", "Search & Discovery", "P2"),
    ("TC053", "test_03_search", "Search page shows Park chip", "Search & Discovery", "P2"),
    ("TC054", "test_03_search", "Search page shows Recent Searches section", "Search & Discovery", "P2"),
    ("TC055", "test_03_search", "/search-results route loads without crashing", "Search & Discovery", "P1"),
    ("TC056", "test_03_search", "Flutter rendered on /search-results route", "Search & Discovery", "P2"),
    ("TC057", "test_03_search", "Empty search state is handled without crashing", "Search & Discovery", "P2"),
    ("TC058", "test_03_search", "Search loading state does not cause a crash", "Search & Discovery", "P2"),
    ("TC059", "test_03_search", "Search page has substantial HTML content", "Search & Discovery", "P2"),
    ("TC060", "test_03_search", "Search page does not show a 500 error", "Search & Discovery", "P1"),
    ("TC061", "test_03_search", "Scrolling on search page does not crash", "Search & Discovery", "P2"),
    ("TC062", "test_03_search", "Browser title is 'crowdsense' on search route", "Search & Discovery", "P2"),
    ("TC063", "test_03_search", "Refreshing search page does not crash", "Search & Discovery", "P2"),
    ("TC064", "test_03_search", "Search page has placeholder text area", "Search & Discovery", "P2"),
    ("TC065", "test_03_search", "Rapid navigation between search and login stable", "Search & Discovery", "P3"),
    # ── Place Details (TC066–TC085) ───────────────────────────────────────────
    ("TC066", "test_04_place", "/place/:id route loads Flutter app", "Place Details", "P1"),
    ("TC067", "test_04_place", "Flutter rendered on place details route", "Place Details", "P1"),
    ("TC068", "test_04_place", "Place details shows loader or content", "Place Details", "P1"),
    ("TC069", "test_04_place", "Place details renders place name area", "Place Details", "P2"),
    ("TC070", "test_04_place", "Place details shows location information", "Place Details", "P2"),
    ("TC071", "test_04_place", "Place details shows a category badge", "Place Details", "P2"),
    ("TC072", "test_04_place", "Place details shows an About section", "Place Details", "P2"),
    ("TC073", "test_04_place", "Place details has Favourite heart icon in app bar", "Place Details", "P2"),
    ("TC074", "test_04_place", "Place details has Share icon in app bar", "Place Details", "P2"),
    ("TC075", "test_04_place", "Place details shows Current Crowd Status card", "Place Details", "P1"),
    ("TC076", "test_04_place", "Place details shows Best Time to Visit section", "Place Details", "P2"),
    ("TC077", "test_04_place", "Place details shows Crowd Forecast section", "Place Details", "P2"),
    ("TC078", "test_04_place", "Place details shows Report Live Crowd Level button", "Place Details", "P1"),
    ("TC079", "test_04_place", "Place details shows SliverAppBar image header", "Place Details", "P2"),
    ("TC080", "test_04_place", "/place/:id/photos loads Community Photos", "Place Details", "P2"),
    ("TC081", "test_04_place", "Invalid place ID does not crash the app", "Place Details", "P1"),
    ("TC082", "test_04_place", "Place details page source has substantial content", "Place Details", "P2"),
    ("TC083", "test_04_place", "Place details page does not show server error", "Place Details", "P1"),
    ("TC084", "test_04_place", "Scrolling on place details page is stable", "Place Details", "P2"),
    ("TC085", "test_04_place", "Browser title is 'crowdsense' on place details route", "Place Details", "P2"),
    # ── Crowd Intelligence (TC086–TC100) ──────────────────────────────────────
    ("TC086", "test_05_crowd", "Place details page renders (crowd test prerequisite)", "Crowd Intelligence", "P1"),
    ("TC087", "test_05_crowd", "Crowd report form shows Low/Not Busy option", "Crowd Intelligence", "P1"),
    ("TC088", "test_05_crowd", "Crowd report form shows Moderate/A bit busy option", "Crowd Intelligence", "P1"),
    ("TC089", "test_05_crowd", "Crowd report form shows High/Very Crowded option", "Crowd Intelligence", "P1"),
    ("TC090", "test_05_crowd", "Crowd report form has optional note field", "Crowd Intelligence", "P2"),
    ("TC091", "test_05_crowd", "Crowd report form has a Submit button", "Crowd Intelligence", "P1"),
    ("TC092", "test_05_crowd", "Place details shows Live Crowd Status card", "Crowd Intelligence", "P1"),
    ("TC093", "test_05_crowd", "Place details shows Best Time to Visit section", "Crowd Intelligence", "P2"),
    ("TC094", "test_05_crowd", "Place details shows Crowd Forecast chart", "Crowd Intelligence", "P2"),
    ("TC095", "test_05_crowd", "Navigating to 3 different place IDs is stable", "Crowd Intelligence", "P2"),
    ("TC096", "test_05_crowd", "Place/crowd page source is not empty", "Crowd Intelligence", "P2"),
    ("TC097", "test_05_crowd", "Crowd page does not show a server error", "Crowd Intelligence", "P1"),
    ("TC098", "test_05_crowd", "Scrolling through crowd details page is stable", "Crowd Intelligence", "P2"),
    ("TC099", "test_05_crowd", "Place details page renders visible content", "Crowd Intelligence", "P2"),
    ("TC100", "test_05_crowd", "Browser back from place details returns safely", "Crowd Intelligence", "P2"),
    # ── Profile (TC101–TC120) ─────────────────────────────────────────────────
    ("TC101", "test_06_profile", "/profile route renders or redirects to login", "Profile & Account", "P1"),
    ("TC102", "test_06_profile", "Unauthenticated user at /profile redirected to login", "Profile & Account", "P1"),
    ("TC103", "test_06_profile", "Profile page shows a heading", "Profile & Account", "P2"),
    ("TC104", "test_06_profile", "Profile page shows Photos stat", "Profile & Account", "P2"),
    ("TC105", "test_06_profile", "Profile page shows Reports stat", "Profile & Account", "P2"),
    ("TC106", "test_06_profile", "Profile page shows Saved stat", "Profile & Account", "P2"),
    ("TC107", "test_06_profile", "Profile stats row has three items", "Profile & Account", "P2"),
    ("TC108", "test_06_profile", "Profile page has Edit Profile menu item", "Profile & Account", "P2"),
    ("TC109", "test_06_profile", "Profile page has Notifications menu item", "Profile & Account", "P2"),
    ("TC110", "test_06_profile", "Profile page has My Reports menu item", "Profile & Account", "P2"),
    ("TC111", "test_06_profile", "Profile page has Support & FAQ menu item", "Profile & Account", "P2"),
    ("TC112", "test_06_profile", "Profile page has Sign Out button", "Profile & Account", "P1"),
    ("TC113", "test_06_profile", "Profile page has Settings icon in app bar", "Profile & Account", "P2"),
    ("TC114", "test_06_profile", "Profile page has camera icon for photo upload", "Profile & Account", "P2"),
    ("TC115", "test_06_profile", "Profile page shows email under user name", "Profile & Account", "P2"),
    ("TC116", "test_06_profile", "Profile page has all required menu items", "Profile & Account", "P2"),
    ("TC117", "test_06_profile", "/my-reports route loads or redirects", "Profile & Account", "P1"),
    ("TC118", "test_06_profile", "Unauthenticated /my-reports redirected to login", "Profile & Account", "P1"),
    ("TC119", "test_06_profile", "My Reports shows list or empty state", "Profile & Account", "P2"),
    ("TC120", "test_06_profile", "Profile page source has substantial content", "Profile & Account", "P2"),
    # ── Favorites (TC121–TC130) ───────────────────────────────────────────────
    ("TC121", "test_07_favorites", "/favorites route loads or redirects gracefully", "Favorites", "P1"),
    ("TC122", "test_07_favorites", "Flutter rendered on /favorites route", "Favorites", "P1"),
    ("TC123", "test_07_favorites", "Favorites accessible for unauthenticated users", "Favorites", "P2"),
    ("TC124", "test_07_favorites", "Favorites shows list or empty state", "Favorites", "P2"),
    ("TC125", "test_07_favorites", "Favorites page renders a heading area", "Favorites", "P2"),
    ("TC126", "test_07_favorites", "Favorites page source has substantial content", "Favorites", "P2"),
    ("TC127", "test_07_favorites", "Favorites page does not show server error", "Favorites", "P1"),
    ("TC128", "test_07_favorites", "Scrolling on favorites page does not crash", "Favorites", "P2"),
    ("TC129", "test_07_favorites", "Browser title is 'crowdsense' on favorites route", "Favorites", "P2"),
    ("TC130", "test_07_favorites", "Refreshing favorites page does not crash", "Favorites", "P2"),
    # ── Travel Planner (TC131–TC145) ──────────────────────────────────────────
    ("TC131", "test_08_planner", "/planner route loads or redirects gracefully", "Travel Planner", "P1"),
    ("TC132", "test_08_planner", "Flutter rendered on /planner route", "Travel Planner", "P1"),
    ("TC133", "test_08_planner", "Travel Planner page shows main heading", "Travel Planner", "P2"),
    ("TC134", "test_08_planner", "Travel Planner shows From location selector", "Travel Planner", "P2"),
    ("TC135", "test_08_planner", "Travel Planner shows To location selector", "Travel Planner", "P2"),
    ("TC136", "test_08_planner", "Travel Planner shows Calculate Best Plan button", "Travel Planner", "P2"),
    ("TC137", "test_08_planner", "Calculate button disabled without From/To selection", "Travel Planner", "P2"),
    ("TC138", "test_08_planner", "Planner shows crowd-aware description text", "Travel Planner", "P2"),
    ("TC139", "test_08_planner", "Travel Planner shows a map section", "Travel Planner", "P2"),
    ("TC140", "test_08_planner", "Travel Planner shows trip summary card area", "Travel Planner", "P2"),
    ("TC141", "test_08_planner", "Planner page source has substantial content", "Travel Planner", "P2"),
    ("TC142", "test_08_planner", "Planner page does not show server error", "Travel Planner", "P1"),
    ("TC143", "test_08_planner", "Scrolling on planner page does not crash", "Travel Planner", "P2"),
    ("TC144", "test_08_planner", "Browser title is 'crowdsense' on planner route", "Travel Planner", "P2"),
    ("TC145", "test_08_planner", "Refreshing planner page does not crash", "Travel Planner", "P2"),
    # ── Settings (TC146–TC165) ────────────────────────────────────────────────
    ("TC146", "test_09_settings", "/settings route loads Flutter app", "Settings", "P1"),
    ("TC147", "test_09_settings", "Settings does NOT require authentication", "Settings", "P1"),
    ("TC148", "test_09_settings", "Flutter rendered on /settings route", "Settings", "P1"),
    ("TC149", "test_09_settings", "Settings page has APPEARANCE section", "Settings", "P2"),
    ("TC150", "test_09_settings", "Settings page has Dark Mode toggle", "Settings", "P2"),
    ("TC151", "test_09_settings", "Settings page has NOTIFICATIONS section", "Settings", "P2"),
    ("TC152", "test_09_settings", "Settings page has Notifications toggle", "Settings", "P2"),
    ("TC153", "test_09_settings", "Settings page has ABOUT section", "Settings", "P2"),
    ("TC154", "test_09_settings", "Settings page shows App Version", "Settings", "P2"),
    ("TC155", "test_09_settings", "Settings page shows version number 1.0.0", "Settings", "P2"),
    ("TC156", "test_09_settings", "Settings page has ACCOUNT section", "Settings", "P2"),
    ("TC157", "test_09_settings", "Settings page has Privacy Policy option", "Settings", "P2"),
    ("TC158", "test_09_settings", "Settings page has Terms of Service option", "Settings", "P2"),
    ("TC159", "test_09_settings", "Settings page has Delete Account option", "Settings", "P2"),
    ("TC160", "test_09_settings", "Settings page source has substantial content", "Settings", "P2"),
    ("TC161", "test_09_settings", "Settings page does not show server error", "Settings", "P1"),
    ("TC162", "test_09_settings", "Scrolling on settings page does not crash", "Settings", "P2"),
    ("TC163", "test_09_settings", "Browser title is 'crowdsense' on settings route", "Settings", "P2"),
    ("TC164", "test_09_settings", "Refreshing settings page does not crash", "Settings", "P2"),
    ("TC165", "test_09_settings", "Browser back from settings returns safely", "Settings", "P2"),
    # ── UI/UX & Responsiveness (TC166–TC185) ──────────────────────────────────
    ("TC166", "test_10_ui_responsive", "App renders correctly on iPhone X (375×812)", "UI/UX & Responsiveness", "P2"),
    ("TC167", "test_10_ui_responsive", "App renders correctly on iPad (768×1024)", "UI/UX & Responsiveness", "P2"),
    ("TC168", "test_10_ui_responsive", "App renders correctly on Full HD (1920×1080)", "UI/UX & Responsiveness", "P2"),
    ("TC169", "test_10_ui_responsive", "App renders correctly on small 320×568 window", "UI/UX & Responsiveness", "P2"),
    ("TC170", "test_10_ui_responsive", "Register page renders on mobile viewport", "UI/UX & Responsiveness", "P2"),
    ("TC171", "test_10_ui_responsive", "Login page has no horizontal scroll overflow", "UI/UX & Responsiveness", "P2"),
    ("TC172", "test_10_ui_responsive", "Browser back/forward navigation works", "UI/UX & Responsiveness", "P2"),
    ("TC173", "test_10_ui_responsive", "Refreshing login page does not crash", "UI/UX & Responsiveness", "P2"),
    ("TC174", "test_10_ui_responsive", "Refreshing register page does not crash", "UI/UX & Responsiveness", "P2"),
    ("TC175", "test_10_ui_responsive", "Rapid navigation through 7 URLs does not crash", "UI/UX & Responsiveness", "P2"),
    ("TC176", "test_10_ui_responsive", "10 consecutive page navigations stay stable", "UI/UX & Responsiveness", "P3"),
    ("TC177", "test_10_ui_responsive", "Scrolling on login page works", "UI/UX & Responsiveness", "P2"),
    ("TC178", "test_10_ui_responsive", "Scrolling on register page works", "UI/UX & Responsiveness", "P2"),
    ("TC179", "test_10_ui_responsive", "Tab key on login does not crash", "UI/UX & Responsiveness", "P3"),
    ("TC180", "test_10_ui_responsive", "App renders on landscape 1024×768", "UI/UX & Responsiveness", "P2"),
    ("TC181", "test_10_ui_responsive", "App page source is valid HTML", "UI/UX & Responsiveness", "P1"),
    ("TC182", "test_10_ui_responsive", "Maximizing after resize restores layout", "UI/UX & Responsiveness", "P2"),
    ("TC183", "test_10_ui_responsive", "Register page has no horizontal overflow", "UI/UX & Responsiveness", "P2"),
    ("TC184", "test_10_ui_responsive", "App not blank after multiple viewport cycles", "UI/UX & Responsiveness", "P2"),
    ("TC185", "test_10_ui_responsive", "App renders complete content on large desktop", "UI/UX & Responsiveness", "P2"),
    # ── Performance (TC186–TC200) ─────────────────────────────────────────────
    ("TC186", "test_11_performance", "App renders visible content within 8 seconds", "Performance", "P1"),
    ("TC187", "test_11_performance", "Login route renders within 8 seconds", "Performance", "P1"),
    ("TC188", "test_11_performance", "manifest.json returns HTTP 200", "Performance", "P2"),
    ("TC189", "test_11_performance", "flutter_bootstrap.js returns HTTP 200", "Performance", "P2"),
    ("TC190", "test_11_performance", "flutter.js returns HTTP 200", "Performance", "P2"),
    ("TC191", "test_11_performance", "favicon.png returns HTTP 200", "Performance", "P2"),
    ("TC192", "test_11_performance", "Base URL returns HTTP 200", "Performance", "P1"),
    ("TC193", "test_11_performance", "Root URL does not show a 404 Not Found", "Performance", "P1"),
    ("TC194", "test_11_performance", "Root URL does not show a 500 Server Error", "Performance", "P1"),
    ("TC195", "test_11_performance", "Page source has adequate length after 5 seconds", "Performance", "P2"),
    ("TC196", "test_11_performance", "Login page source has adequate length", "Performance", "P2"),
    ("TC197", "test_11_performance", "Register page source has adequate length", "Performance", "P2"),
    ("TC198", "test_11_performance", "Forgot password page source has adequate length", "Performance", "P2"),
    ("TC199", "test_11_performance", "App responds to route changes within 10 seconds", "Performance", "P2"),
    ("TC200", "test_11_performance", "Settings page loads within 6 seconds", "Performance", "P2"),
    # ── Edge Cases (TC201–TC215) ──────────────────────────────────────────────
    ("TC201", "test_12_edge", "Unknown route shows fallback or redirects gracefully", "Edge Cases", "P2"),
    ("TC202", "test_12_edge", "Deeply nested invalid URL does not crash", "Edge Cases", "P2"),
    ("TC203", "test_12_edge", "All protected routes redirect unauthenticated user", "Edge Cases", "P1"),
    ("TC204", "test_12_edge", "Auth routes accessible without session", "Edge Cases", "P1"),
    ("TC205", "test_12_edge", "Invalid place ID does not crash the app", "Edge Cases", "P2"),
    ("TC206", "test_12_edge", "Empty hash route (#/) handled without crashing", "Edge Cases", "P2"),
    ("TC207", "test_12_edge", "Rapid navigation through 15 URLs does not crash", "Edge Cases", "P3"),
    ("TC208", "test_12_edge", "10 consecutive navigations keep app stable", "Edge Cases", "P3"),
    ("TC209", "test_12_edge", "5 Tab key presses on login page do not crash", "Edge Cases", "P3"),
    ("TC210", "test_12_edge", "Page source valid after cycling all auth routes", "Edge Cases", "P2"),
    ("TC211", "test_12_edge", "/home route properly guards unauthenticated users", "Edge Cases", "P1"),
    ("TC212", "test_12_edge", "/profile route properly guards unauthenticated users", "Edge Cases", "P1"),
    ("TC213", "test_12_edge", "/my-reports route properly guards unauthenticated users", "Edge Cases", "P1"),
    ("TC214", "test_12_edge", "App stable after window resize cycle", "Edge Cases", "P2"),
    ("TC215", "test_12_edge", "Rapid search/place navigation is stable", "Edge Cases", "P2"),
    # ── Accessibility (TC216–TC225) ───────────────────────────────────────────
    ("TC216", "test_13_accessibility", "Tab key on login page does not crash", "Accessibility", "P3"),
    ("TC217", "test_13_accessibility", "Tab key on register page does not crash", "Accessibility", "P3"),
    ("TC218", "test_13_accessibility", "Tab key on forgot-password page does not crash", "Accessibility", "P3"),
    ("TC219", "test_13_accessibility", "Page title is not empty", "Accessibility", "P2"),
    ("TC220", "test_13_accessibility", "Page title contains 'crowdsense'", "Accessibility", "P2"),
    ("TC221", "test_13_accessibility", "HTML element has a lang attribute", "Accessibility", "P3"),
    ("TC222", "test_13_accessibility", "Scroll position resets on route change", "Accessibility", "P3"),
    ("TC223", "test_13_accessibility", "Enter key on login page does not crash", "Accessibility", "P3"),
    ("TC224", "test_13_accessibility", "Focus not trapped after navigating from login", "Accessibility", "P3"),
    ("TC225", "test_13_accessibility", "No autoscroll loop after loading login page", "Accessibility", "P3"),
    # ── Smoke Tests (TC226–TC240) ─────────────────────────────────────────────
    ("TC226", "test_14_smoke", "SMOKE: Base URL loads the Flutter app shell", "Smoke Tests", "P1"),
    ("TC227", "test_14_smoke", "SMOKE: Browser title contains 'crowdsense'", "Smoke Tests", "P1"),
    ("TC228", "test_14_smoke", "SMOKE: Flutter CanvasKit rendered on base URL", "Smoke Tests", "P1"),
    ("TC229", "test_14_smoke", "SMOKE: /login route loads", "Smoke Tests", "P1"),
    ("TC230", "test_14_smoke", "SMOKE: /register route loads", "Smoke Tests", "P1"),
    ("TC231", "test_14_smoke", "SMOKE: /forgot-password route loads", "Smoke Tests", "P1"),
    ("TC232", "test_14_smoke", "SMOKE: /settings route loads", "Smoke Tests", "P1"),
    ("TC233", "test_14_smoke", "SMOKE: /search route loads", "Smoke Tests", "P1"),
    ("TC234", "test_14_smoke", "SMOKE: /place/:id route loads without crashing", "Smoke Tests", "P1"),
    ("TC235", "test_14_smoke", "SMOKE: /home either loads or redirects", "Smoke Tests", "P1"),
    ("TC236", "test_14_smoke", "SMOKE: /profile redirects unauthenticated to login", "Smoke Tests", "P1"),
    ("TC237", "test_14_smoke", "SMOKE: manifest.json is accessible", "Smoke Tests", "P1"),
    ("TC238", "test_14_smoke", "SMOKE: flutter_bootstrap.js is accessible", "Smoke Tests", "P1"),
    ("TC239", "test_14_smoke", "SMOKE: Root URL does not return 404", "Smoke Tests", "P1"),
    ("TC240", "test_14_smoke", "SMOKE: Root URL does not return 500", "Smoke Tests", "P1"),
    # ── Vulnerability Tests (TC241–TC340) ─────────────────────────────────────
    ("TC241", "test_16_vulnerability", "Verify HTTPS enforcement and HTTP redirection on root URL", "Vulnerability Tests", "P1"),
    ("TC242", "test_16_vulnerability", "Verify modern SSL/TLS version support (TLS 1.2 or 1.3)", "Vulnerability Tests", "P1"),
    ("TC243", "test_16_vulnerability", "Verify Strict-Transport-Security (HSTS) header presence", "Vulnerability Tests", "P1"),
    ("TC244", "test_16_vulnerability", "Verify HTTP request to root redirects to secure origin", "Vulnerability Tests", "P1"),
    ("TC245", "test_16_vulnerability", "Verify SSL certificate is within its valid date range", "Vulnerability Tests", "P1"),
    ("TC246", "test_16_vulnerability", "Verify SSL certificate is signed by a recognized public CA", "Vulnerability Tests", "P1"),
    ("TC247", "test_16_vulnerability", "Verify SSL certificate hostname matches the target domain", "Vulnerability Tests", "P1"),
    ("TC248", "test_16_vulnerability", "Verify secure TLS cipher suite configuration", "Vulnerability Tests", "P2"),
    ("TC249", "test_16_vulnerability", "Verify no active mixed content (HTTP assets on HTTPS page)", "Vulnerability Tests", "P1"),
    ("TC250", "test_16_vulnerability", "Verify TLS session resumption does not leak tickets", "Vulnerability Tests", "P3"),
    ("TC251", "test_16_vulnerability", "Verify Content-Security-Policy (CSP) header is present", "Vulnerability Tests", "P1"),
    ("TC252", "test_16_vulnerability", "Verify CSP script-src directive restricts unsafe-inline", "Vulnerability Tests", "P1"),
    ("TC253", "test_16_vulnerability", "Verify X-Frame-Options is set to DENY or SAMEORIGIN", "Vulnerability Tests", "P1"),
    ("TC254", "test_16_vulnerability", "Verify X-Content-Type-Options is set to nosniff", "Vulnerability Tests", "P1"),
    ("TC255", "test_16_vulnerability", "Verify Referrer-Policy header is configured securely", "Vulnerability Tests", "P2"),
    ("TC256", "test_16_vulnerability", "Verify Permissions-Policy restricts unused APIs", "Vulnerability Tests", "P2"),
    ("TC257", "test_16_vulnerability", "Verify Server header does not disclose software versions", "Vulnerability Tests", "P2"),
    ("TC258", "test_16_vulnerability", "Verify X-Powered-By header is omitted from responses", "Vulnerability Tests", "P2"),
    ("TC259", "test_16_vulnerability", "Verify X-XSS-Protection header is present or configured", "Vulnerability Tests", "P2"),
    ("TC260", "test_16_vulnerability", "Verify Cache-Control restricts storing private user state", "Vulnerability Tests", "P2"),
    ("TC261", "test_16_vulnerability", "Verify robots.txt does not disclose sensitive paths", "Vulnerability Tests", "P2"),
    ("TC262", "test_16_vulnerability", "Verify .git repository folder returns 404 or is blocked", "Vulnerability Tests", "P1"),
    ("TC263", "test_16_vulnerability", "Verify common configuration backup files return 404", "Vulnerability Tests", "P2"),
    ("TC264", "test_16_vulnerability", "Verify JavaScript source map files are not exposed", "Vulnerability Tests", "P2"),
    ("TC265", "test_16_vulnerability", "Verify package.json is not exposed in production root", "Vulnerability Tests", "P2"),
    ("TC266", "test_16_vulnerability", "Verify pubspec.yaml/lock is not exposed in production root", "Vulnerability Tests", "P2"),
    ("TC267", "test_16_vulnerability", "Verify HTML comments contain no sensitive credentials or IPs", "Vulnerability Tests", "P2"),
    ("TC268", "test_16_vulnerability", "Verify backup files (.bak, .tmp, ~) return 404", "Vulnerability Tests", "P2"),
    ("TC269", "test_16_vulnerability", "Verify developer READMEs are not exposed in deployment root", "Vulnerability Tests", "P2"),
    ("TC270", "test_16_vulnerability", "Verify directory listing/indexing is disabled", "Vulnerability Tests", "P2"),
    ("TC271", "test_16_vulnerability", "Verify Access-Control-Allow-Origin does not allow wildcard with credentials", "Vulnerability Tests", "P1"),
    ("TC272", "test_16_vulnerability", "Verify CORS preflight methods are properly restricted", "Vulnerability Tests", "P2"),
    ("TC273", "test_16_vulnerability", "Verify Cross-Origin-Embedder-Policy (COEP) presence", "Vulnerability Tests", "P3"),
    ("TC274", "test_16_vulnerability", "Verify Cross-Origin-Opener-Policy (COOP) presence", "Vulnerability Tests", "P3"),
    ("TC275", "test_16_vulnerability", "Verify Cross-Origin-Resource-Policy (CORP) presence", "Vulnerability Tests", "P3"),
    ("TC276", "test_16_vulnerability", "Verify iframe injection restriction from external sources", "Vulnerability Tests", "P2"),
    ("TC277", "test_16_vulnerability", "Verify no routing redirect vulnerabilities to untrusted hosts", "Vulnerability Tests", "P2"),
    ("TC278", "test_16_vulnerability", "Verify target=_blank links use rel=noopener/noreferrer", "Vulnerability Tests", "P2"),
    ("TC279", "test_16_vulnerability", "Verify Access-Control-Max-Age is set to limit preflights", "Vulnerability Tests", "P3"),
    ("TC280", "test_16_vulnerability", "Verify CORS wildcard responses reject authorization headers", "Vulnerability Tests", "P2"),
    ("TC281", "test_16_vulnerability", "Verify autocomplete attribute is present on login username/email", "Vulnerability Tests", "P2"),
    ("TC282", "test_16_vulnerability", "Verify login password field uses type=password masking", "Vulnerability Tests", "P1"),
    ("TC283", "test_16_vulnerability", "Verify register password field uses type=password masking", "Vulnerability Tests", "P1"),
    ("TC284", "test_16_vulnerability", "Verify register confirm password field uses type=password", "Vulnerability Tests", "P1"),
    ("TC285", "test_16_vulnerability", "Verify sensitive fields disable general browser autofill", "Vulnerability Tests", "P2"),
    ("TC286", "test_16_vulnerability", "Verify plaintext passwords are not leaked in DOM attributes", "Vulnerability Tests", "P1"),
    ("TC287", "test_16_vulnerability", "Verify input character limits prevent client-side overflow", "Vulnerability Tests", "P2"),
    ("TC288", "test_16_vulnerability", "Verify forms submit via POST instead of exposing data in GET", "Vulnerability Tests", "P1"),
    ("TC289", "test_16_vulnerability", "Verify standard frontend verification controls are configured", "Vulnerability Tests", "P2"),
    ("TC290", "test_16_vulnerability", "Verify click events do not leak credentials in raw parameters", "Vulnerability Tests", "P2"),
    ("TC291", "test_16_vulnerability", "Verify /home guards unauthenticated session redirects to login", "Vulnerability Tests", "P1"),
    ("TC292", "test_16_vulnerability", "Verify /profile guards unauthenticated session redirects to login", "Vulnerability Tests", "P1"),
    ("TC293", "test_16_vulnerability", "Verify /favorites guards unauthenticated session redirects to login", "Vulnerability Tests", "P1"),
    ("TC294", "test_16_vulnerability", "Verify /planner guards unauthenticated session redirects to login", "Vulnerability Tests", "P1"),
    ("TC295", "test_16_vulnerability", "Verify /settings guards or handles route state safely", "Vulnerability Tests", "P1"),
    ("TC296", "test_16_vulnerability", "Verify /admin or invalid routes redirect or return 404", "Vulnerability Tests", "P1"),
    ("TC297", "test_16_vulnerability", "Verify deeply nested unauthorized links redirect to login", "Vulnerability Tests", "P2"),
    ("TC298", "test_16_vulnerability", "Verify route history does not expose authorization credentials", "Vulnerability Tests", "P2"),
    ("TC299", "test_16_vulnerability", "Verify sensitive state variables are cleared on route changes", "Vulnerability Tests", "P2"),
    ("TC300", "test_16_vulnerability", "Verify logout clears route token permissions", "Vulnerability Tests", "P1"),
    ("TC301", "test_16_vulnerability", "Verify local storage contains no plaintext passwords", "Vulnerability Tests", "P1"),
    ("TC302", "test_16_vulnerability", "Verify session storage does not expose unencrypted PII", "Vulnerability Tests", "P2"),
    ("TC303", "test_16_vulnerability", "Verify IndexedDB is empty of plaintext authentication secret keys", "Vulnerability Tests", "P2"),
    ("TC304", "test_16_vulnerability", "Verify cookies set by application contain Secure attribute", "Vulnerability Tests", "P1"),
    ("TC305", "test_16_vulnerability", "Verify cookies set by application contain HttpOnly attribute", "Vulnerability Tests", "P1"),
    ("TC306", "test_16_vulnerability", "Verify cookies set by application enforce SameSite attribute", "Vulnerability Tests", "P2"),
    ("TC307", "test_16_vulnerability", "Verify browser cache metadata prevents public caching of token requests", "Vulnerability Tests", "P2"),
    ("TC308", "test_16_vulnerability", "Verify Clear-Site-Data is used during logout flow", "Vulnerability Tests", "P3"),
    ("TC309", "test_16_vulnerability", "Verify authentication tokens in storage have expiration attributes", "Vulnerability Tests", "P2"),
    ("TC310", "test_16_vulnerability", "Verify backend database configuration details are not local", "Vulnerability Tests", "P1"),
    ("TC311", "test_16_vulnerability", "Verify subresource integrity (SRI) hashes on CDN scripts", "Vulnerability Tests", "P2"),
    ("TC312", "test_16_vulnerability", "Verify inline event handlers are not dynamically written", "Vulnerability Tests", "P2"),
    ("TC313", "test_16_vulnerability", "Verify eval() calls are absent from runtime scripts", "Vulnerability Tests", "P2"),
    ("TC314", "test_16_vulnerability", "Verify document.write() is not utilized in dynamic scripts", "Vulnerability Tests", "P2"),
    ("TC315", "test_16_vulnerability", "Verify HTML templates use safe bindings instead of raw innerHTML", "Vulnerability Tests", "P2"),
    ("TC316", "test_16_vulnerability", "Verify third-party script references use secure HTTPS protocol", "Vulnerability Tests", "P1"),
    ("TC317", "test_16_vulnerability", "Verify window.name contains no active session or auth token data", "Vulnerability Tests", "P3"),
    ("TC318", "test_16_vulnerability", "Verify client-side error handling escapes user-controlled inputs", "Vulnerability Tests", "P2"),
    ("TC319", "test_16_vulnerability", "Verify dynamic script additions are restricted to trust sources", "Vulnerability Tests", "P2"),
    ("TC320", "test_16_vulnerability", "Verify textContent is preferred over innerHTML for dynamic labels", "Vulnerability Tests", "P2"),
    ("TC321", "test_16_vulnerability", "Verify session identifiers have sufficient entropy (>128 bits)", "Vulnerability Tests", "P1"),
    ("TC322", "test_16_vulnerability", "Verify credentials are not auto-populated in forms prior to user interaction", "Vulnerability Tests", "P2"),
    ("TC323", "test_16_vulnerability", "Verify rapid login rate-limiting responses are handled gracefully", "Vulnerability Tests", "P2"),
    ("TC324", "test_16_vulnerability", "Verify credentials and verification codes are not in code bundle", "Vulnerability Tests", "P1"),
    ("TC325", "test_16_vulnerability", "Verify JWT token validation rejects 'none' algorithm client-side", "Vulnerability Tests", "P1"),
    ("TC326", "test_16_vulnerability", "Verify profile URL parameter does not leak sequential guessing", "Vulnerability Tests", "P2"),
    ("TC327", "test_16_vulnerability", "Verify API calls incorporate standard Authorization headers", "Vulnerability Tests", "P1"),
    ("TC328", "test_16_vulnerability", "Verify token expiration is detected and triggers auth workflow reset", "Vulnerability Tests", "P2"),
    ("TC329", "test_16_vulnerability", "Verify change password flow includes validation of the old password", "Vulnerability Tests", "P2"),
    ("TC330", "test_16_vulnerability", "Verify auth error messages do not disclose username/email existence", "Vulnerability Tests", "P2"),
    ("TC331", "test_16_vulnerability", "Verify manifest.json contains a secure relative start_url", "Vulnerability Tests", "P2"),
    ("TC332", "test_16_vulnerability", "Verify manifest.json structure leaks no developer environment details", "Vulnerability Tests", "P2"),
    ("TC333", "test_16_vulnerability", "Verify default CMS path triggers (wp-admin, dashboard.php) return 404", "Vulnerability Tests", "P2"),
    ("TC334", "test_16_vulnerability", "Verify geolocation or device permissions are not implicitly requested", "Vulnerability Tests", "P2"),
    ("TC335", "test_16_vulnerability", "Verify telemetry or metric scripts do not intercept sensitive input fields", "Vulnerability Tests", "P2"),
    ("TC336", "test_16_vulnerability", "Verify console.log does not output user credentials or session tokens", "Vulnerability Tests", "P2"),
    ("TC337", "test_16_vulnerability", "Verify stylesheets contain no references to untrusted image domains", "Vulnerability Tests", "P2"),
    ("TC338", "test_16_vulnerability", "Verify service worker is registered securely from matching origin", "Vulnerability Tests", "P2"),
    ("TC339", "test_16_vulnerability", "Verify client-side router does not permit wildcard open redirects", "Vulnerability Tests", "P1"),
    ("TC340", "test_16_vulnerability", "Verify security configuration checks are successfully passed", "Vulnerability Tests", "P1"),
]


def generate_report(output_path: str | None = None) -> str:
    """Generate the XLSX test report template with all 340 test cases."""
    if output_path is None:
        ts = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
        reports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
        os.makedirs(reports_dir, exist_ok=True)
        output_path = os.path.join(reports_dir, f"CrowdSense_E2E_Template_{ts}.xlsx")

    wb = openpyxl.Workbook()
    total = len(TEST_CASES)
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    _build_summary(ws1, total, now_str)

    # ── Sheet 2: Test Results (full table) ────────────────────────────────────
    ws2 = wb.create_sheet("Test Results")
    _build_results(ws2, total, now_str)

    # ── Sheet 3: Test Scenarios (by category) ─────────────────────────────────
    ws3 = wb.create_sheet("Test Scenarios")
    _build_scenarios(ws3)

    wb.save(output_path)
    print(f"\n[OK] Report template saved → {output_path}")
    return output_path


def _build_summary(ws, total: int, now_str: str):
    # Title
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = "CrowdSense — End-to-End Selenium Test Report Template"
    t.font      = _font(bold=True, size=18)
    t.fill      = HDR_FILL
    t.alignment = CENTER
    ws.row_dimensions[1].height = 50

    # Subtitle
    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value     = (f"Application: CrowdSense Web App  |  "
                   f"URL: {BASE_URL}  |  "
                   f"Framework: Selenium 4 + pytest  |  Generated: {now_str}")
    s.font      = _font(size=9)
    s.fill      = _fill(_BLUE)
    s.alignment = CENTER
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 12  # spacer

    # Stats header
    stat_labels = ["Total Tests", "✅ Passed", "❌ Failed", "⚠️ Pending", "Pass Rate", "Duration"]
    stat_values = [total, 0, 0, total, "0.00%", "—"]
    stat_fills  = [_fill(_BLUE), PASS_FILL, FAIL_FILL, PEND_FILL, _fill("D9E2F3"), _fill("D9E2F3")]

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 38
    for col_idx, (label, val, fill) in enumerate(zip(stat_labels, stat_values, stat_fills), start=1):
        lc = ws.cell(row=4, column=col_idx, value=label)
        lc.font = _font(color="000000", bold=True, size=10)
        lc.fill = fill; lc.alignment = CENTER; lc.border = THIN_BDR
        vc = ws.cell(row=5, column=col_idx, value=val)
        vc.font = _font(color="000000", bold=True, size=14)
        vc.fill = fill; vc.alignment = CENTER; vc.border = THIN_BDR

    ws.row_dimensions[6].height = 12  # spacer
    info = [
        ("Application",    "CrowdSense"),
        ("Type",           "Flutter Web App"),
        ("Framework",      "Selenium 4 + pytest"),
        ("Environment",    "Production — GitHub Pages"),
        ("Base URL",       BASE_URL),
        ("Total Tests",    str(total)),
        ("Passed",         "0"),
        ("Failed",         "0"),
        ("Pending",        str(total)),
        ("Pass Rate",      "0.00%"),
        ("Prepared By",    "CrowdSense Automated Test Suite"),
        ("Report Version", "v2.0 — Template"),
    ]
    _hdr(ws, ["Property", "Value"], row=7)
    for r, (k, v) in enumerate(info, start=8):
        for col, val in [(1, k), (2, v)]:
            c = ws.cell(row=r, column=col, value=val)
            c.font = BOLD if col == 1 else NORMAL
            c.border = THIN_BDR; c.alignment = LEFT
            if r % 2 == 0:
                c.fill = ALT_FILL
        ws.row_dimensions[r].height = 18

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 10


def _build_results(ws, total: int, now_str: str):
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "CrowdSense — E2E Test Cases & Results"
    t.font = _font(bold=True, size=14); t.fill = HDR_FILL; t.alignment = CENTER
    ws.row_dimensions[1].height = 36

    headers = ["TC ID", "Module", "Test Case Description", "Category",
               "Priority", "Status", "Actual Result / Notes", "Duration (s)", "Timestamp"]
    widths  = [10, 28, 70, 24, 12, 12, 55, 16, 22]
    _hdr(ws, headers, row=2)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 28

    for r_idx, (tc_id, module, desc, cat, prio) in enumerate(TEST_CASES, start=3):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        vals = [tc_id, module, desc, cat, prio, "PENDING", "— Not yet executed —", "", now_str]
        for c_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = NORMAL; c.border = THIN_BDR; c.alignment = LEFT
            if fill:
                c.fill = fill
        st_cell = ws.cell(row=r_idx, column=6)
        st_cell.fill  = PEND_FILL
        st_cell.font  = _font(color="9C6500", bold=True, size=10)
        st_cell.alignment = CENTER
        ws.row_dimensions[r_idx].height = 20

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:I{total + 2}"


def _build_scenarios(ws):
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "CrowdSense — Test Scenario Coverage"
    t.font = _font(bold=True, size=14); t.fill = HDR_FILL; t.alignment = CENTER
    ws.row_dimensions[1].height = 36

    scenarios = [
        ("Authentication",         25, "Auth flows, route guards, page elements"),
        ("Home & Navigation",       20, "Home route, categories, nav bar, sections"),
        ("Search & Discovery",      20, "Search input, chips, results, recent"),
        ("Place Details",           20, "Place info, icons, community photos"),
        ("Crowd Intelligence",      15, "Crowd report form, status card, forecast"),
        ("Profile & Account",       20, "Profile stats, menus, my reports"),
        ("Favorites",               10, "Favorites route, list/empty state"),
        ("Travel Planner",          15, "Planner selectors, map, calculate button"),
        ("Settings",                20, "Dark mode, notifications, version, sections"),
        ("UI/UX & Responsiveness",  20, "Viewports, overflow, back/forward, scroll"),
        ("Performance",             15, "Load times, static asset HTTP status"),
        ("Edge Cases",              15, "Unknown routes, stress, guards"),
        ("Accessibility",           10, "Keyboard, tab, focus, title, scroll"),
        ("Smoke Tests",             15, "Critical path sanity — all main routes"),
        ("Vulnerability Tests",     100, "Transport layer, headers, disclosure, CORS, forms, guards, script safety"),
    ]

    _hdr(ws, ["Feature Area", "Test Cases", "Description", "Priority", "Status"], row=2)
    for r_idx, (area, count, desc) in enumerate(scenarios, start=3):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, val in enumerate([area, count, desc, "P1–P3", "PENDING"], start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = NORMAL; c.border = THIN_BDR; c.alignment = LEFT
            if fill:
                c.fill = fill
        ws.row_dimensions[r_idx].height = 20

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14


def _hdr(ws, columns, row=1):
    for col_idx, hdr in enumerate(columns, start=1):
        c = ws.cell(row=row, column=col_idx, value=hdr)
        c.font = WHITE_HDR; c.fill = HDR_FILL
        c.border = THIN_BDR; c.alignment = CENTER
    ws.row_dimensions[row].height = 28


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate a CrowdSense E2E test report template (XLSX)"
    )
    parser.add_argument("--output", "-o", help="Output file path (optional)")
    args = parser.parse_args()

    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    path = generate_report(args.output)
    print(f"\nOpen the report at: {path}")
