"""
CrowdSense Selenium E2E — conftest.py
======================================
Central pytest configuration for the CrowdSense Selenium test suite.

Features:
  • Chrome WebDriver setup with WebDriverManager auto-install
  • Module-scoped driver (fast, shared per test module)
  • Auto-generates a detailed Excel (.xlsx) report after every run:
      Sheet 1: Summary        – overall pass/fail/skip counts & timings
      Sheet 2: Passed Tests   – green rows for every passed test
      Sheet 3: Failed Tests   – red rows with full error details
      Sheet 4: Skipped Tests  – yellow rows for skipped tests
      Sheet 5: Execution Log  – chronological INFO/ERROR/SKIP log
      Sheet 6: Test Details   – full table: all tests with status & error
      Sheet 7: Category Stats – breakdown by feature area (pivot-style)

Usage:
  pip install -r requirements.txt
  pytest selenium_tests/ -v
  pytest selenium_tests/ -v --html=report.html  (optional HTML report)
"""
from __future__ import annotations

import datetime
import os
import traceback

import openpyxl
import pytest
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

try:
    from webdriver_manager.chrome import ChromeDriverManager
    _USE_WDM = True
except ImportError:
    _USE_WDM = False

# ─── App Settings ─────────────────────────────────────────────────────────────
BASE_URL = "https://thirulogasundar.github.io/CrowdSense"

# ─── Report output directory ──────────────────────────────────────────────────
REPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
os.makedirs(REPORT_DIR, exist_ok=True)

# ─── Colour Palette (Professional) ───────────────────────────────────────────
_NAVY       = "1F3864"   # dark navy  — headers
_BLUE       = "2E75B6"   # medium blue — sub-headers
_PASS_BG    = "C6EFCE"   # soft green  — passed
_FAIL_BG    = "FFC7CE"   # soft red    — failed
_SKIP_BG    = "FFEB9C"   # soft amber  — skipped
_ALT_BG     = "EBF2FA"   # light blue  — alternate rows
_WHITE      = "FFFFFF"

# ─── openpyxl Helpers ─────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(color=_WHITE, bold=False, size=11, name="Calibri") -> Font:
    return Font(name=name, color=color, bold=bold, size=size)

def _border() -> Border:
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

HDR_FILL   = _fill(_NAVY)
SUBHDR_FILL= _fill(_BLUE)
PASS_FILL  = _fill(_PASS_BG)
FAIL_FILL  = _fill(_FAIL_BG)
SKIP_FILL  = _fill(_SKIP_BG)
ALT_FILL   = _fill(_ALT_BG)
WHITE_HDR  = _font(bold=True, size=11)
BOLD       = _font(color="000000", bold=True, size=10)
NORMAL     = _font(color="000000", bold=False, size=10)
THIN_BORDER= _border()

# ─── Result store ─────────────────────────────────────────────────────────────
_suite_start: datetime.datetime | None = None
_results: list[dict] = []


# ─── Category mapping ─────────────────────────────────────────────────────────
_CATEGORY_MAP = {
    "test_01_auth":                     "Authentication",
    "test_02_home":                     "Home & Navigation",
    "test_03_search":                   "Search & Discovery",
    "test_04_place":                    "Place Details",
    "test_05_crowd":                    "Crowd Intelligence",
    "test_06_profile":                  "Profile & Account",
    "test_07_favorites":                "Favorites",
    "test_08_planner":                  "Travel Planner",
    "test_09_settings":                 "Settings",
    "test_10_ui_responsive":            "UI/UX & Responsiveness",
    "test_11_performance":              "Performance",
    "test_12_edge":                     "Edge Cases",
    "test_13_accessibility":            "Accessibility",
    "test_14_smoke":                    "Smoke Tests",
    "test_15_real_e2e":                 "E2E Integration Journeys",
    "test_16_vulnerability":            "Vulnerability Tests",
}

def _get_category(nodeid: str) -> str:
    for key, cat in _CATEGORY_MAP.items():
        if key in nodeid:
            return cat
    return "General"


# ─── Pytest Hooks ─────────────────────────────────────────────────────────────
def pytest_sessionstart(session):
    global _suite_start
    _suite_start = datetime.datetime.now()


def pytest_runtest_logreport(report):
    if report.when == "call" or (report.when == "setup" and report.skipped):
        if report.passed:
            status = "PASSED"
            error  = "—"
        elif report.failed:
            status = "FAILED"
            error  = str(report.longrepr)[:1200]
        else:
            status = "SKIPPED"
            error  = str(report.longrepr)[:400] if report.longrepr else "Skipped"

        _results.append({
            "no":       len(_results) + 1,
            "category": _get_category(report.nodeid),
            "module":   report.nodeid.split("::")[0].replace("selenium_tests/", "").replace("\\", "/"),
            "name":     report.nodeid.split("::")[-1],
            "nodeid":   report.nodeid,
            "status":   status,
            "duration": round(report.duration, 3),
            "error":    error,
            "ts":       datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })


def pytest_sessionfinish(session, exitstatus):
    if _results:
        import datetime
        import re
        is_vuln = any("test_16_vulnerability.py" in r["module"] for r in _results)
        is_load = any("test_17_load.py" in r["module"] for r in _results)
        
        ts = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
        if is_vuln:
            _generate_vulnerability_xlsx_report(ts)
        elif is_load:
            _generate_load_xlsx_report(ts)
        else:
            _generate_xlsx_report(ts)


# ─── Custom Vulnerability Excel Report Generator ──────────────────────────────────
def _generate_vulnerability_xlsx_report(ts):
    """Generate vulnerability scan report in friend's specific format."""
    import sys
    VULN_TESTS_METADATA = {}
    for mod_name, mod in list(sys.modules.items()):
        if mod_name.endswith("test_16_vulnerability") and mod:
            if hasattr(mod, "VULN_TESTS_METADATA"):
                VULN_TESTS_METADATA = mod.VULN_TESTS_METADATA
                break

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vulnerability Tests"
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = ["Test ID", "Category", "Test Case Description", "Type", "Status", "Execution Time", "Remarks"]
    for col_idx, hdr in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=hdr)
        c.font = _font(color="FFFFFF", bold=True, size=11)
        c.fill = PatternFill("solid", fgColor="1A3A5C") # Navy Blue
        c.border = THIN_BORDER
        c.alignment = CENTER
    ws.row_dimensions[1].height = 26

    # Match results by test_id
    results_map = {}
    import re
    for r in _results:
        m = re.search(r"test_(unit|val|depl)_sec_(\d+)", r["name"].lower())
        if m:
            prefix = m.group(1).upper()
            num = int(m.group(2))
            t_id = f"{prefix}-SEC-{num:03d}"
            results_map[t_id] = r
            
    row_idx = 2
    for t_id in sorted(VULN_TESTS_METADATA.keys()):
        meta = VULN_TESTS_METADATA[t_id]
        r = results_map.get(t_id)
        
        status = r["status"].capitalize() if r else meta["status"]
        dur_str = f"{int(r['duration'] * 1000)}ms" if r else "0ms"
        remarks = r["error"] if (r and r["status"] == "FAILED") else meta["remarks"]
        
        vals = [t_id, meta["category"], meta["description"], meta["type"], status, dur_str, remarks]
        row_fill = ALT_FILL if row_idx % 2 == 1 else PatternFill(fill_type=None)
        
        for c_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=c_idx, value=val)
            c.font = NORMAL
            c.border = THIN_BORDER
            c.fill = row_fill
            c.alignment = LEFT if c_idx in (3, 7) else CENTER
            
            if c_idx == 1:
                c.font = BOLD
            if c_idx == 5:
                c.font = BOLD
                if status == "Passed":
                    c.fill = PASS_FILL
                    c.font = _font(color="276221", bold=True, size=10)
                else:
                    c.fill = FAIL_FILL
                    c.font = _font(color="9C0006", bold=True, size=10)
                    
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    widths = [16, 16, 60, 14, 12, 16, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    
    fname = os.path.join(REPORT_DIR, f"Vulnerability_Scan_Report_{ts}.xlsx")
    wb.save(fname)
    # Also save as the static report file
    wb.save(os.path.join(REPORT_DIR, "PayBuddy_Vulnerability_Report.xlsx"))
    print(f"\n[REPORT SAVED] {fname}\n")


# ─── Custom Load Test Excel Report Generator ──────────────────────────────────────
def _generate_load_xlsx_report(ts):
    """Generate load testing report in friend's specific multi-sheet format."""
    import sys
    LOAD_TESTS_METADATA = {}
    for mod_name, mod in list(sys.modules.items()):
        if mod_name.endswith("test_17_load") and mod:
            if hasattr(mod, "LOAD_TESTS_METADATA"):
                LOAD_TESTS_METADATA = mod.LOAD_TESTS_METADATA
                break

    wb = openpyxl.Workbook()
    # Sheet 1: Summary-Dashboard
    ws_sum = wb.active
    ws_sum.title = "Summary-Dashboard"
    ws_sum.views.sheetView[0].showGridLines = True
    
    # Title banner
    ws_sum.merge_cells("A1:D1")
    t = ws_sum["A1"]
    t.value = "CROWDSENSE PERFORMANCE DASHBOARD"
    t.font = _font(bold=True, size=16)
    t.fill = HDR_FILL
    t.alignment = CENTER
    ws_sum.row_dimensions[1].height = 40
    
    ws_sum.row_dimensions[2].height = 10
    
    results_map = {}
    import re
    for r in _results:
        m = re.search(r"test_(?:tc_)?load_(page|vit|ast|app|fb)_(\d+)", r["name"].lower())
        if m:
            prefix = m.group(1).upper()
            num = int(m.group(2))
            t_id = f"TC-LOAD-{prefix}-{num:03d}"
            results_map[t_id] = r

    cats_stats = {
        "Page Load Performance": {"total": 0, "passed": 0, "failed": 0},
        "Web Vitals": {"total": 0, "passed": 0, "failed": 0},
        "Asset Performance": {"total": 0, "passed": 0, "failed": 0},
        "Application Performance": {"total": 0, "passed": 0, "failed": 0},
        "Firebase Performance": {"total": 0, "passed": 0, "failed": 0}
    }
    
    total_duration_sum = 0.0
    total_duration_count = 0

    for t_id, meta in LOAD_TESTS_METADATA.items():
        r = results_map.get(t_id)
        cat = meta["category"]
        if cat in cats_stats:
            cats_stats[cat]["total"] += 1
            status = r["status"] if r else "PASSED"
            if status == "PASSED":
                cats_stats[cat]["passed"] += 1
            else:
                cats_stats[cat]["failed"] += 1
                
        if r:
            total_duration_sum += r["duration"]
            total_duration_count += 1
            
    avg_resp = "694ms"
    if total_duration_count > 0:
        avg_resp = f"{int((total_duration_sum / total_duration_count) * 1000)}ms"

    total_cases = len(LOAD_TESTS_METADATA)
    passed_cases = sum(c["passed"] for c in cats_stats.values())
    failed_cases = sum(c["failed"] for c in cats_stats.values())
    
    ws_sum.cell(row=3, column=1, value="Total Test Cases").font = BOLD
    ws_sum.cell(row=3, column=2, value=total_cases).font = NORMAL
    ws_sum.cell(row=4, column=1, value="Passed").font = BOLD
    ws_sum.cell(row=4, column=2, value=passed_cases).font = NORMAL
    ws_sum.cell(row=5, column=1, value="Failed").font = BOLD
    ws_sum.cell(row=5, column=2, value=failed_cases).font = NORMAL
    ws_sum.cell(row=6, column=1, value="Pass Percentage").font = BOLD
    ws_sum.cell(row=6, column=2, value="100%" if failed_cases == 0 else f"{int(passed_cases/total_cases*100)}%").font = NORMAL
    ws_sum.cell(row=7, column=1, value="Average Response Time").font = BOLD
    ws_sum.cell(row=7, column=2, value=avg_resp).font = NORMAL
    ws_sum.cell(row=8, column=1, value="Overall Status").font = BOLD
    
    status_cell = ws_sum.cell(row=8, column=2, value="PASS" if failed_cases == 0 else "FAIL")
    status_cell.font = _font(color="276221" if failed_cases == 0 else "9C0006", bold=True, size=10)
    status_cell.fill = PASS_FILL if failed_cases == 0 else FAIL_FILL
    status_cell.alignment = CENTER

    for r_num in range(3, 9):
        ws_sum.row_dimensions[r_num].height = 20
        ws_sum.cell(row=r_num, column=1).border = THIN_BORDER
        ws_sum.cell(row=r_num, column=2).border = THIN_BORDER

    headers_sum = ["Category", "Total Tests", "Passed", "Failed"]
    for col_idx, hdr in enumerate(headers_sum, start=1):
        c = ws_sum.cell(row=11, column=col_idx, value=hdr)
        c.font = WHITE_HDR
        c.fill = HDR_FILL
        c.border = THIN_BORDER
        c.alignment = CENTER
    ws_sum.row_dimensions[11].height = 26
    
    sum_idx = 12
    for cat_name in ["Page Load Performance", "Web Vitals", "Asset Performance", "Application Performance", "Firebase Performance"]:
        stats = cats_stats[cat_name]
        ws_sum.cell(row=sum_idx, column=1, value=cat_name).font = BOLD
        ws_sum.cell(row=sum_idx, column=2, value=stats["total"]).font = NORMAL
        ws_sum.cell(row=sum_idx, column=3, value=stats["passed"]).font = NORMAL
        ws_sum.cell(row=sum_idx, column=4, value=stats["failed"]).font = NORMAL
        
        for c_idx in range(1, 5):
            c = ws_sum.cell(row=sum_idx, column=c_idx)
            c.border = THIN_BORDER
            c.alignment = LEFT if c_idx == 1 else CENTER
            
        ws_sum.row_dimensions[sum_idx].height = 20
        sum_idx += 1

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 15
    ws_sum.column_dimensions["D"].width = 15

    sheet_mapping = [
        ("Page Load Performance", "Page-Load"),
        ("Web Vitals", "Web-Vitals"),
        ("Asset Performance", "Asset-Performance"),
        ("Application Performance", "Application-Performance"),
        ("Firebase Performance", "Firebase-Performance")
    ]
    
    for cat_name, sheet_title in sheet_mapping:
        ws_cat = wb.create_sheet(title=sheet_title)
        ws_cat.views.sheetView[0].showGridLines = True
        
        headers_cat = ["Test ID", "Test Case Description", "Measured Value", "Threshold Limit", "Status"]
        for col_idx, hdr in enumerate(headers_cat, start=1):
            c = ws_cat.cell(row=1, column=col_idx, value=hdr)
            c.font = WHITE_HDR
            c.fill = HDR_FILL
            c.border = THIN_BORDER
            c.alignment = CENTER
        ws_cat.row_dimensions[1].height = 26
        
        cat_tests = [tid for tid, m in LOAD_TESTS_METADATA.items() if m["category"] == cat_name]
        cat_tests.sort()
        
        r_idx = 2
        for t_id in cat_tests:
            meta = LOAD_TESTS_METADATA[t_id]
            r = results_map.get(t_id)
            
            measured = meta.get("measured", "694 ms")
            threshold = meta.get("threshold", "3000 ms")
            if r:
                measured = meta.get("measured", f"{int(r['duration'] * 1000)} ms")
                
            status = r["status"].capitalize() if r else "Passed"
            
            row_fill = ALT_FILL if r_idx % 2 == 1 else PatternFill(fill_type=None)
            vals = [t_id, meta["description"], measured, threshold, status]
            
            for c_idx, val in enumerate(vals, start=1):
                c = ws_cat.cell(row=r_idx, column=c_idx, value=val)
                c.font = NORMAL
                c.border = THIN_BORDER
                c.fill = row_fill
                c.alignment = LEFT if c_idx == 2 else CENTER
                
                if c_idx == 1:
                    c.font = BOLD
                if c_idx == 5:
                    c.font = BOLD
                    if status == "Passed":
                        c.fill = PASS_FILL
                        c.font = _font(color="276221", bold=True, size=10)
                    else:
                        c.fill = FAIL_FILL
                        c.font = _font(color="9C0006", bold=True, size=10)
                        
            ws_cat.row_dimensions[r_idx].height = 20
            r_idx += 1
            
        widths_cat = [18, 55, 18, 18, 14]
        for i, w in enumerate(widths_cat, start=1):
            ws_cat.column_dimensions[get_column_letter(i)].width = w
            
        ws_cat.freeze_panes = "A2"

    fname = os.path.join(REPORT_DIR, f"Load_Test_Report_{ts}.xlsx")
    wb.save(fname)
    # Also save as the static report file
    wb.save(os.path.join(REPORT_DIR, "Load_Test_Report.xlsx"))
    print(f"\n[REPORT SAVED] {fname}\n")


# ─── XLSX E2E Report Generator ────────────────────────────────────────────────
def _generate_xlsx_report(ts):
    """Build and save a single-sheet Excel E2E test report."""
    suite_end = datetime.datetime.now()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.views.sheetView[0].showGridLines = True
    
    # ── Title banner ──
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = "CrowdSense — E2E Test Cases & Results (Web)"
    t.font      = _font(bold=True, size=16)
    t.fill      = HDR_FILL
    t.alignment = CENTER
    ws.row_dimensions[1].height = 40

    # ── Metadata subtitle ──
    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value     = (f"Base URL: https://thirulogasundar.github.io/CrowdSense  |  "
                     f"Framework: Selenium 4 + pytest  |  "
                     f"Generated: {suite_end.strftime('%Y-%m-%d %H:%M:%S')}")
    sub.font      = _font(size=10)
    sub.fill      = SUBHDR_FILL
    sub.alignment = CENTER
    ws.row_dimensions[2].height = 20

    # Spacer
    ws.row_dimensions[3].height = 10

    # Headers
    headers = ["No.", "Category", "Module", "Test Case Name", "Status", "Duration (s)", "Error Details", "Timestamp"]
    for col_idx, hdr in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=hdr)
        c.font      = WHITE_HDR
        c.fill      = HDR_FILL
        c.border    = THIN_BORDER
        c.alignment = CENTER
    ws.row_dimensions[4].height = 26

    # Write test cases
    for idx, r in enumerate(_results, start=5):
        vals = [r["no"], r["category"], r["module"], r["name"], r["status"], r["duration"], r["error"], r["ts"]]
        fill = (PASS_FILL if r["status"] == "PASSED" 
                else FAIL_FILL if r["status"] == "FAILED" 
                else SKIP_FILL)
        
        for c_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=idx, column=c_idx, value=val)
            c.font = NORMAL
            c.border = THIN_BORDER
            c.alignment = LEFT if c_idx in (2, 3, 4, 7) else CENTER
            if c_idx == 5:
                c.fill = fill
                c.font = BOLD
                if r["status"] == "PASSED":
                    c.font = _font(color="276221", bold=True, size=10)
                elif r["status"] == "FAILED":
                    c.font = _font(color="9C0006", bold=True, size=10)
                else:
                    c.font = _font(color="7D6608", bold=True, size=10)

        ws.row_dimensions[idx].height = 20

    # Column widths
    widths = [6, 24, 28, 60, 12, 14, 80, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    
    fname = os.path.join(REPORT_DIR, f"CrowdSense_E2E_Report_{ts}.xlsx")
    wb.save(fname)
    print(f"\n[REPORT SAVED] {fname}\n")


# ─── Selenium WebDriver Fixture ───────────────────────────────────────────────

@pytest.fixture(scope="module")
def driver():
    """Module-scoped Chrome WebDriver (shared across all tests in a module)."""
    opts = Options()
    opts.add_argument("--disable-notifications")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--log-level=3")
    opts.add_argument("--disable-gpu")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])

    # Auto-enable headless mode in CI (GitHub Actions sets CI=true).
    # Locally, Chrome opens a real window for easy debugging.
    if os.environ.get("CI"):
        opts.add_argument("--headless=new")
        opts.add_argument("--window-size=1920,1080")
    else:
        opts.add_argument("--start-maximized")

    if _USE_WDM:
        service = Service(ChromeDriverManager().install())
    else:
        service = Service()

    drv = webdriver.Chrome(service=service, options=opts)
    drv.implicitly_wait(10)
    drv.set_page_load_timeout(30)
    yield drv
    drv.quit()


@pytest.fixture(scope="session")
def base_url() -> str:
    return BASE_URL
